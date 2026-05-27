from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime, timezone
import io
import logging

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import db
from models import UserRole
from auth_utils import (
    get_current_user, require_roles, calculate_grade, create_audit_log, get_teacher_assigned_classes
)

router = APIRouter()
logger = logging.getLogger(__name__)


# ==================== DASHBOARD ====================

@router.get("/reports/dashboard")
async def get_dashboard_stats(request: Request, academic_year: Optional[str] = None):
    user = await get_current_user(request)
    stats = {}

    if user["role"] in [UserRole.ADMIN, UserRole.ACCOUNTANT]:
        # Session scoping: when an academic_year is selected, student & fee
        # figures reflect that session only. Employees/issues are global.
        stu_q = {"is_active": True}
        if academic_year:
            stu_q["academic_year"] = academic_year
        stats["total_students"] = await db.students.count_documents(stu_q)
        stats["total_employees"] = await db.employees.count_documents({"is_active": True})
        stats["fee_overdue_count"] = await db.students.count_documents(
            {**stu_q, "fee_status": "overdue"}
        )
        stats["open_issues"] = await db.issues.count_documents({"status": "open"})

        # "Today's present" must reflect the selected session, not the live
        # clock — otherwise every session shows the same number. Use real today
        # for the active session (whose window extends to today); for a past
        # session there is no "today", so anchor to its last day.
        today = datetime.now().strftime("%Y-%m-%d")
        ref_date = today
        if academic_year:
            active = await db.sessions.find_one({"is_active": True}, {"_id": 0, "session_name": 1})
            active_ay = active.get("session_name") if active else None
            try:
                sy = int(academic_year.split("-")[0])
                s_start, s_end = f"{sy}-04-01", f"{sy + 1}-03-31"
                if academic_year == active_ay:
                    ref_date = today if today >= s_start else s_start
                else:
                    ref_date = today if s_start <= today <= s_end else s_end
            except (ValueError, IndexError):
                ref_date = today
        present_count = await db.attendance.count_documents({
            "entity_type": "student", "date": ref_date, "status": "present"
        })
        stats["today_present"] = present_count

        # For a selected session, show that session's total collection; for the
        # live/active session keep the current-month figure.
        if academic_year:
            coll_agg = await db.fee_payments.aggregate([
                {"$match": {"academic_year": academic_year}},
                {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
            ]).to_list(1)
            stats["month_collection"] = coll_agg[0]["total"] if coll_agg else 0
        else:
            current_month = datetime.now().strftime("%Y-%m")
            month_agg = await db.fee_payments.aggregate([
                {"$match": {"payment_date": {"$regex": f"^{current_month}"}}},
                {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
            ]).to_list(1)
            stats["month_collection"] = month_agg[0]["total"] if month_agg else 0

    if user["role"] == UserRole.TEACHER:
        today = datetime.now().strftime("%Y-%m-%d")
        teacher_attendance = await db.attendance.find(
            {"marked_by": user["user_id"]}, {"_id": 0, "class_name": 1, "section": 1}
        ).to_list(10000)
        unique_classes = set()
        for a in teacher_attendance:
            if a.get("class_name") and a.get("section"):
                unique_classes.add(f"{a['class_name']}-{a['section']}")
        stats["assigned_classes"] = len(unique_classes)
        today_marked = await db.attendance.count_documents({"marked_by": user["user_id"], "date": today})
        stats["pending_attendance"] = max(len(unique_classes) - (1 if today_marked > 0 else 0), 0)
        stats["pending_marks_entry"] = 0

    if user["role"] == UserRole.STUDENT:
        student = await db.students.find_one({"user_id": user["user_id"]}, {"_id": 0})
        if student:
            stats["fee_status"] = student["fee_status"]
            stats["app_locked"] = student.get("app_locked", False)
            stats["class_name"] = student.get("class_name", "")
            stats["section"] = student.get("section", "")
            stats["roll_number"] = student.get("roll_number", "")
            stats["admission_number"] = student.get("admission_number", "")
            stats["stream"] = student.get("stream", "")
            attendance = await db.attendance.find({
                "entity_type": "student", "entity_id": student["student_id"]
            }, {"_id": 0}).to_list(1000)
            if attendance:
                present = sum(1 for a in attendance if a["status"] == "present")
                stats["attendance_percentage"] = round(present / len(attendance) * 100, 1)
            else:
                stats["attendance_percentage"] = 0

    if user["role"] == UserRole.PARENT:
        children = await db.students.find(
            {"$or": [{"parent_email": user.get("email", "")}, {"parent_id": user["user_id"]}], "is_active": True},
            {"_id": 0}
        ).to_list(10)
        stats["children_count"] = len(children)
        stats["children"] = children
        # Check if any child is app_locked
        stats["app_locked"] = any(c.get("app_locked", False) for c in children)

    return stats


# ==================== FINANCIAL REPORT ====================

@router.get("/reports/financial")
async def get_financial_report(request: Request, start_date: Optional[str] = None, end_date: Optional[str] = None, academic_year: Optional[str] = None):
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    import asyncio as _asyncio

    # Scope every figure to the selected session so each academic year shows
    # only its own collection/pending totals (not the same global numbers).
    match = {}
    if academic_year:
        match["academic_year"] = academic_year
    if start_date and end_date:
        match["payment_date"] = {"$gte": start_date, "$lte": end_date}

    pending_match = {"status": {"$in": ["pending", "overdue"]}}
    if academic_year:
        pending_match["academic_year"] = academic_year

    # All aggregations run in parallel — no full collection scan needed
    total_agg, method_agg, month_agg, pending_agg, count = await _asyncio.gather(
        db.fee_payments.aggregate([
            {"$match": match},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
        ]).to_list(1),
        db.fee_payments.aggregate([
            {"$match": match},
            {"$group": {"_id": "$payment_method", "total": {"$sum": "$amount"}}}
        ]).to_list(20),
        db.fee_payments.aggregate([
            {"$match": match},
            {"$group": {"_id": {"$substr": ["$payment_date", 0, 7]}, "total": {"$sum": "$amount"}}},
            {"$sort": {"_id": 1}},
        ]).to_list(36),
        db.student_ledger.aggregate([
            {"$match": pending_match},
            {"$group": {"_id": None, "total": {"$sum": "$net_amount"}}}
        ]).to_list(1),
        db.fee_payments.count_documents(match),
    )

    return {
        "total_collection": total_agg[0]["total"] if total_agg else 0,
        "total_pending": pending_agg[0]["total"] if pending_agg else 0,
        "by_payment_method": {r["_id"] or "unknown": r["total"] for r in method_agg},
        "by_month": {r["_id"]: r["total"] for r in month_agg if r["_id"]},
        "transaction_count": count,
    }


@router.get("/reports/financial/export")
async def export_financial_report(request: Request, format: str = "pdf", start_date: Optional[str] = None, end_date: Optional[str] = None, academic_year: Optional[str] = None):
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)

    query = {}
    if academic_year:
        query["academic_year"] = academic_year
    if start_date and end_date:
        query["payment_date"] = {"$gte": start_date, "$lte": end_date}

    pending_match = {"status": {"$in": ["pending", "overdue"]}}
    if academic_year:
        pending_match["academic_year"] = academic_year

    import asyncio as _asyncio
    payments, total_agg, pending_agg = await _asyncio.gather(
        db.fee_payments.find(query, {"_id": 0}).sort("payment_date", -1).limit(2000).to_list(2000),
        db.fee_payments.aggregate([
            {"$match": query},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
        ]).to_list(1),
        db.student_ledger.aggregate([
            {"$match": pending_match},
            {"$group": {"_id": None, "total": {"$sum": "$net_amount"}}}
        ]).to_list(1),
    )
    total_collection = total_agg[0]["total"] if total_agg else 0
    total_pending = pending_agg[0]["total"] if pending_agg else 0

    if format == "excel":
        return _financial_excel(payments, total_collection, total_pending)
    else:
        return _financial_pdf(payments, total_collection, total_pending)


def _financial_pdf(payments, total_collection, total_pending):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER, spaceAfter=20)
    elements = []

    elements.append(Paragraph("Shemford Futuristic School", title_style))
    elements.append(Paragraph("Financial Report", ParagraphStyle('Sub', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=20)))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}", styles['Normal']))
    elements.append(Spacer(1, 12))

    # Summary
    summary_data = [
        ['Metric', 'Value'],
        ['Total Collection', f'Rs. {total_collection:,.2f}'],
        ['Total Pending', f'Rs. {total_pending:,.2f}'],
        ['Transactions', str(len(payments))],
    ]
    t = Table(summary_data, colWidths=[200, 200])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E88A1A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 20))

    # Payment Details
    if payments:
        elements.append(Paragraph("Payment Details", styles['Heading3']))
        elements.append(Spacer(1, 8))
        table_data = [['Receipt', 'Student ID', 'Amount', 'Method', 'Month', 'Date']]
        for p in payments[:100]:
            table_data.append([
                p.get('receipt_number', '-')[:15],
                p.get('student_id', '-')[:15],
                f'Rs. {p["amount"]:,.0f}',
                p.get('payment_method', '-'),
                p.get('month', '-'),
                p.get('payment_date', '-'),
            ])
        t2 = Table(table_data, colWidths=[85, 90, 70, 60, 60, 70])
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#333333')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        elements.append(t2)

    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=financial_report.pdf"}
    )


def _financial_excel(payments, total_collection, total_pending):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    header_fill = PatternFill(start_color="E88A1A", end_color="E88A1A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "Shemford Futuristic School - Financial Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ws.append([])

    # Summary
    ws.append(["Summary"])
    ws.append(["Total Collection", f"Rs. {total_collection:,.2f}"])
    ws.append(["Total Pending", f"Rs. {total_pending:,.2f}"])
    ws.append(["Transactions", len(payments)])
    ws.append([])

    # Headers
    headers = ['Receipt No.', 'Student ID', 'Amount', 'Method', 'Month', 'Date']
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    # Data
    for p in payments:
        ws.append([
            p.get('receipt_number', ''),
            p.get('student_id', ''),
            p['amount'],
            p.get('payment_method', ''),
            p.get('month', ''),
            p.get('payment_date', ''),
        ])

    # Auto width
    for col in ws.columns:
        try:
            col_letter = col[0].column_letter
            max_length = max(len(str(cell.value or '')) for cell in col if cell.value is not None)
            ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
        except AttributeError:
            pass

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=financial_report.xlsx"}
    )


# ==================== ATTENDANCE REPORT ====================

@router.get("/reports/attendance")
async def get_attendance_report(
    request: Request,
    class_name: Optional[str] = None,
    date: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    user = await require_roles(UserRole.ADMIN, UserRole.TEACHER)(request)
    query = {"entity_type": "student"}

    # Teachers can only see their assigned class's attendance report
    if user["role"] == UserRole.TEACHER:
        assigned = await get_teacher_assigned_classes(user["user_id"])
        if assigned:
            allowed_classes = [a["class_name"] for a in assigned]
            if class_name and class_name not in allowed_classes:
                raise HTTPException(status_code=403, detail="You are not assigned to this class")
            if not class_name:
                query["$or"] = [{"class_name": a["class_name"], "section": a["section"]} for a in assigned]

    if class_name:
        query["class_name"] = class_name
    if date:
        query["date"] = date
    elif start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["date"] = {"$gte": start_date}
    elif end_date:
        query["date"] = {"$lte": end_date}

    import asyncio as _asyncio

    # Run all queries in parallel — counts done by MongoDB, only 200 rows fetched for display
    total, present, absent, late, records = await _asyncio.gather(
        db.attendance.count_documents(query),
        db.attendance.count_documents({**query, "status": "present"}),
        db.attendance.count_documents({**query, "status": "absent"}),
        db.attendance.count_documents({**query, "status": "late"}),
        db.attendance.find(query, {"_id": 0}).sort("date", -1).limit(200).to_list(200),
    )

    return {
        "total_records": total,
        "present": present,
        "absent": absent,
        "late": late,
        "percentage": round((present / total * 100), 1) if total > 0 else 0,
        "records": records,
    }


@router.get("/reports/attendance/export")
async def export_attendance_report(
    request: Request,
    format: str = "pdf",
    class_name: Optional[str] = None,
    date: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    user = await require_roles(UserRole.ADMIN, UserRole.TEACHER)(request)
    query = {"entity_type": "student"}
    if user["role"] == UserRole.TEACHER:
        assigned = await get_teacher_assigned_classes(user["user_id"])
        if assigned:
            allowed_classes = [a["class_name"] for a in assigned]
            if class_name and class_name not in allowed_classes:
                raise HTTPException(status_code=403, detail="You are not assigned to this class")
            if not class_name:
                query["$or"] = [{"class_name": a["class_name"], "section": a["section"]} for a in assigned]
    if class_name:
        query["class_name"] = class_name
    if date:
        query["date"] = date
    elif start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["date"] = {"$gte": start_date}
    elif end_date:
        query["date"] = {"$lte": end_date}

    records = await db.attendance.find(query, {"_id": 0}).to_list(10000)

    # Batch-fetch all student names in one query instead of N queries
    student_ids = list({r["entity_id"] for r in records})
    students_list = await db.students.find(
        {"student_id": {"$in": student_ids}},
        {"_id": 0, "student_id": 1, "first_name": 1, "last_name": 1, "admission_number": 1}
    ).to_list(len(student_ids))
    student_map = {s["student_id"]: s for s in students_list}

    for r in records:
        student = student_map.get(r["entity_id"])
        if student:
            r["student_name"] = f"{student['first_name']} {student['last_name']}"
            r["admission_number"] = student.get("admission_number", "")
        else:
            r["student_name"] = r["entity_id"]
            r["admission_number"] = ""

    if format == "excel":
        return _attendance_excel(records, class_name, date)
    else:
        return _attendance_pdf(records, class_name, date)


def _attendance_pdf(records, class_name, date):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Shemford Futuristic School", ParagraphStyle('T', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER)))
    subtitle = "Attendance Report"
    if class_name:
        subtitle += f" - Class {class_name}"
    if date:
        subtitle += f" - {date}"
    elements.append(Paragraph(subtitle, ParagraphStyle('S', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=20)))

    # Summary
    total = len(records)
    present = sum(1 for r in records if r["status"] == "present")
    absent = sum(1 for r in records if r["status"] == "absent")
    pct = round((present / total * 100), 1) if total > 0 else 0
    elements.append(Paragraph(f"Total: {total} | Present: {present} | Absent: {absent} | Attendance: {pct}%", styles['Normal']))
    elements.append(Spacer(1, 12))

    if records:
        table_data = [['Adm. No.', 'Student Name', 'Class', 'Section', 'Date', 'Status']]
        for r in records[:200]:
            table_data.append([
                r.get('admission_number', '')[:12],
                r.get('student_name', '')[:25],
                r.get('class_name', ''),
                r.get('section', ''),
                r.get('date', ''),
                r.get('status', '').upper(),
            ])
        t = Table(table_data, colWidths=[70, 130, 50, 50, 70, 60])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E88A1A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=attendance_report.pdf"})


def _attendance_excel(records, class_name, date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    header_fill = PatternFill(start_color="E88A1A", end_color="E88A1A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws.merge_cells('A1:F1')
    ws['A1'] = f"Attendance Report - {f'Class {class_name}' if class_name else 'All'} - {date or 'All Dates'}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    headers = ['Admission No.', 'Student Name', 'Class', 'Section', 'Date', 'Status']
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    for r in records:
        ws.append([
            r.get('admission_number', ''),
            r.get('student_name', ''),
            r.get('class_name', ''),
            r.get('section', ''),
            r.get('date', ''),
            r.get('status', ''),
        ])

    for col in ws.columns:
        try:
            col_letter = col[0].column_letter
            max_length = max(len(str(cell.value or '')) for cell in col if cell.value is not None)
            ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
        except AttributeError:
            pass

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=attendance_report.xlsx"})


# ==================== ACADEMIC REPORT ====================

@router.get("/reports/academic")
async def get_academic_report(request: Request, class_name: Optional[str] = None, section: Optional[str] = None, academic_year: str = "2024-2025"):
    await require_roles(UserRole.ADMIN, UserRole.TEACHER)(request)

    query = {"academic_year": academic_year}
    if class_name:
        query["class_name"] = class_name
    if section:
        query["section"] = section

    marks = await db.mark_records.find(query, {"_id": 0}).to_list(10000)

    student_results = {}
    for m in marks:
        sid = m["student_id"]
        if sid not in student_results:
            student_results[sid] = {"marks": [], "total_obtained": 0, "total_max": 0}
        student_results[sid]["marks"].append(m)
        student_results[sid]["total_obtained"] += m["marks_obtained"]
        student_results[sid]["total_max"] += m["max_marks"]

    class_average = 0
    if student_results:
        percentages = []
        for data in student_results.values():
            if data["total_max"] > 0:
                pct = (data["total_obtained"] / data["total_max"]) * 100
                percentages.append(pct)
                data["percentage"] = round(pct, 2)
                data["grade"] = calculate_grade(pct)
        if percentages:
            class_average = round(sum(percentages) / len(percentages), 2)

    return {
        "academic_year": academic_year,
        "class_name": class_name,
        "section": section,
        "student_count": len(student_results),
        "class_average": class_average,
        "student_results": student_results
    }


@router.get("/reports/academic/export")
async def export_academic_report(request: Request, format: str = "pdf", class_name: Optional[str] = None, section: Optional[str] = None, academic_year: str = "2024-2025"):
    await require_roles(UserRole.ADMIN, UserRole.TEACHER)(request)

    query = {"academic_year": academic_year}
    if class_name:
        query["class_name"] = class_name
    if section:
        query["section"] = section

    marks = await db.mark_records.find(query, {"_id": 0}).to_list(10000)

    # Aggregate per student
    student_results = {}
    for m in marks:
        sid = m["student_id"]
        if sid not in student_results:
            student_results[sid] = {"marks": [], "total_obtained": 0, "total_max": 0}
        student_results[sid]["marks"].append(m)
        student_results[sid]["total_obtained"] += m["marks_obtained"]
        student_results[sid]["total_max"] += m["max_marks"]

    # Enrich with names
    for sid in student_results:
        student = await db.students.find_one({"student_id": sid}, {"_id": 0, "first_name": 1, "last_name": 1, "admission_number": 1})
        if student:
            student_results[sid]["name"] = f"{student['first_name']} {student['last_name']}"
            student_results[sid]["admission_number"] = student.get("admission_number", "")
        else:
            student_results[sid]["name"] = sid
            student_results[sid]["admission_number"] = ""
        if student_results[sid]["total_max"] > 0:
            pct = (student_results[sid]["total_obtained"] / student_results[sid]["total_max"]) * 100
            student_results[sid]["percentage"] = round(pct, 2)
            student_results[sid]["grade"] = calculate_grade(pct)

    if format == "excel":
        return _academic_excel(student_results, class_name, section, academic_year)
    else:
        return _academic_pdf(student_results, class_name, section, academic_year)


def _academic_pdf(student_results, class_name, section, academic_year):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Shemford Futuristic School", ParagraphStyle('T', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER)))
    subtitle = f"Academic Report - {academic_year}"
    if class_name:
        subtitle += f" | Class {class_name}"
    if section:
        subtitle += f" - {section}"
    elements.append(Paragraph(subtitle, ParagraphStyle('S', parent=styles['Heading2'], alignment=TA_CENTER, spaceAfter=20)))

    if student_results:
        table_data = [['Adm. No.', 'Student Name', 'Marks', 'Total', 'Percentage', 'Grade']]
        for sid, data in student_results.items():
            table_data.append([
                data.get('admission_number', '')[:12],
                data.get('name', '')[:25],
                f"{data['total_obtained']:.0f}",
                f"{data['total_max']:.0f}",
                f"{data.get('percentage', 0):.1f}%",
                data.get('grade', '-'),
            ])
        t = Table(table_data, colWidths=[75, 150, 55, 55, 60, 40])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E88A1A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("No academic records found for the selected criteria.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=academic_report.pdf"})


def _academic_excel(student_results, class_name, section, academic_year):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Academic Report"

    header_fill = PatternFill(start_color="E88A1A", end_color="E88A1A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws.merge_cells('A1:F1')
    ws['A1'] = f"Academic Report - {academic_year} - {f'Class {class_name}' if class_name else 'All'}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    headers = ['Admission No.', 'Student Name', 'Marks Obtained', 'Total Marks', 'Percentage', 'Grade']
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    for data in student_results.values():
        ws.append([
            data.get('admission_number', ''),
            data.get('name', ''),
            data['total_obtained'],
            data['total_max'],
            f"{data.get('percentage', 0):.1f}%",
            data.get('grade', '-'),
        ])

    for col in ws.columns:
        try:
            col_letter = col[0].column_letter
            max_length = max(len(str(cell.value or '')) for cell in col if cell.value is not None)
            ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
        except AttributeError:
            pass

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=academic_report.xlsx"})


# ==================== REMINDERS ====================

@router.post("/reports/reminders")
async def send_fee_reminders(request: Request):
    user = await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)  # noqa: F841
    overdue_students = await db.students.find({"fee_status": {"$in": ["pending", "overdue"]}}, {"_id": 0}).to_list(2000)

    results = {"sent": 0, "failed": 0, "details": []}

    for student in overdue_students:
        email = student.get("parent_email") or student.get("email")
        if email:
            results["details"].append({
                "student": f"{student['first_name']} {student['last_name']}",
                "email": email,
                "status": "queued"
            })
            results["sent"] += 1

    return results


# ==================== AUDIT LOGS ====================

@router.get("/audit-logs")
async def get_audit_logs(request: Request, entity_type: Optional[str] = None, entity_id: Optional[str] = None, limit: int = 50):
    await require_roles(UserRole.ADMIN)(request)
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    if entity_id:
        query["entity_id"] = entity_id

    logs = await db.audit_logs.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return logs


# ==================== HEALTH CHECK ====================

@router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}
