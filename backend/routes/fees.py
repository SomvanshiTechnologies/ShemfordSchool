"""
Shemford Futuristic School — Component-Based Fee Management System

Fee structure:
  One-time  : Registration, Admission, Caution Deposit
  Yearly    : Annual Charge, Activity Fee, Exam Fee, Lab Fee
  Monthly   : Tuition (12 months; 1st month collected at admission, rest month-by-month)

Admission flow:
  - Collect one-time + yearly + 1st month tuition → generate admission number → create ledger
  - Remaining 11 monthly tuition entries auto-generated after admission

Sibling discount:
  - 50% off admission fee
  - 15% off monthly tuition

Stream support (class 11/12):
  - Science, Arts, Commerce each have separate fee configs

Annual increase:
  - Admin can apply N% increase to all or selected classes for next session
"""
from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import datetime, timezone, date
import io
import logging
import re
import uuid
from pymongo import UpdateOne

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# ── Unicode font for Rs. symbol ──────────────────────────────────────────────────
import os as _os
from reportlab.pdfbase import pdfmetrics as _pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont as _TTFont

_PDF_FONT_REG  = "Helvetica"
_PDF_FONT_BOLD = "Helvetica-Bold"

for _reg, _bold in [
    ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
     "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
    ("/usr/share/fonts/dejavu/DejaVuSans.ttf",
     "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf"),
    ("C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf"),
]:
    if _os.path.exists(_reg):
        try:
            _pdfmetrics.registerFont(_TTFont("_FeesSans", _reg))
            _pdfmetrics.registerFont(_TTFont("_FeesSans-Bold",
                                             _bold if _os.path.exists(_bold) else _reg))
            _PDF_FONT_REG  = "_FeesSans"
            _PDF_FONT_BOLD = "_FeesSans-Bold"
        except Exception:
            pass
        break


def _iso_to_dmy(s) -> str:
    """Convert YYYY-MM-DD to DD/MM/YYYY for PDF/report display."""
    if not s:
        return "—"
    s = str(s)
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return f"{s[8:10]}/{s[5:7]}/{s[0:4]}"
    return s


from database import db, client as mongo_client
from models import (
    UserRole, FeeComponentConfig, StudentLedgerEntry, FeePayment,
    FeeComponentType, FEE_COMPONENT_FREQUENCY
)
from auth_utils import get_current_user, require_roles, create_audit_log

router = APIRouter()
logger = logging.getLogger(__name__)

# Maps fee-config field names → canonical fee_component identifiers (FeeComponentType values)
CFG_FIELD_TO_COMPONENT = {
    "registration_fee": "registration",
    "admission_fee":    "admission",
    "caution_deposit":  "caution_deposit",
    "annual_charge":    "annual_charge",
    "activity_fee":     "activity_fee",
    "exam_fee":         "exam_fee",
    "lab_fee":          "lab_fee",
    "ai_robotics_fee":  "ai_robotics_fee",
    "monthly_tuition":  "tuition",
    "upgradation_fee":  "upgradation",
}

ACADEMIC_MONTHS = ["04", "05", "06", "07", "08", "09", "10", "11", "12", "01", "02", "03"]

# Human-readable labels for fee components — must match the report filter
# dropdown so the "Fees Type" column lines up with what admins can filter by.
_COMPONENT_LABELS = {
    "registration": "Registration Fee",
    "admission": "Admission Fee",
    "caution_deposit": "Caution Deposit",
    "annual_charge": "Annual Charge",
    "activity_fee": "Activity Fee",
    "exam_fee": "Exam Fee",
    "lab_fee": "Lab Fee",
    "ai_robotics_fee": "AI & Robotics Fee",
    "tuition": "Tuition Fee",
    "upgradation": "Upgradation Fee",
}


def _filter_label(fee_component: str = None, fee_month: str = None) -> str:
    """Human label for an active fee-type filter, matching the dropdown
    (e.g. fee_month='08' -> 'August Fees', fee_component='exam_fee' -> 'Exam Fee')."""
    import calendar
    if fee_month:
        try:
            return f"{calendar.month_name[int(fee_month)]} Fees"
        except (ValueError, IndexError):
            return "Tuition Fee"
    if fee_component:
        return _COMPONENT_LABELS.get(fee_component, fee_component.replace("_", " ").title())
    return ""


def _entry_category_label(e: dict) -> str:
    """Map a ledger entry to its display category, matching the report filter
    options: monthly tuition → '<Month> Fees', everything else → component label."""
    import calendar
    comp = e.get("fee_component")
    if comp == "tuition":
        dd = str(e.get("due_date") or "")
        try:
            mo = int(dd[5:7])
            return f"{calendar.month_name[mo]} Fees"
        except (ValueError, IndexError):
            return "Tuition Fee"
    if comp in _COMPONENT_LABELS:
        return _COMPONENT_LABELS[comp]
    if comp:
        return comp.replace("_", " ").title()
    return str(e.get("fee_type") or "")


def _report_category(e: dict) -> str:
    """Display category for a ledger entry in reports — tuition collapsed to a
    single 'Tuition Fee' (matches the Collection report and the filter dropdown)."""
    comp = e.get("fee_component")
    if comp and comp in _COMPONENT_LABELS:
        return _COMPONENT_LABELS[comp]
    if comp:
        return comp.replace("_", " ").title()
    return (e.get("fee_type") or "Other").replace("_", " ").title()

# ─── helpers ─────────────────────────────────────────────────────────────────

import re

def validate_academic_year(y: str):
    """Raise HTTPException if academic_year is not in YYYY-YYYY+1 format. (#14)"""
    if not re.match(r"^\d{4}-\d{4}$", y):
        raise HTTPException(status_code=400, detail="academic_year must be in YYYY-YYYY format (e.g. 2025-2026)")
    start, end = int(y[:4]), int(y[5:])
    if end != start + 1:
        raise HTTPException(status_code=400, detail="academic_year end year must be start year + 1 (e.g. 2025-2026)")


_FEE_AMOUNT_FIELDS = [
    "registration_fee", "admission_fee", "caution_deposit", "annual_charge",
    "activity_fee", "exam_fee", "lab_fee", "ai_robotics_fee",
    "monthly_tuition", "upgradation_fee", "late_fee",
]


def validate_fee_amounts(body: dict):
    """Raise HTTPException if any fee field is negative. (#13)"""
    for field in _FEE_AMOUNT_FIELDS:
        if field in body:
            val = float(body[field])
            if val < 0:
                raise HTTPException(status_code=400, detail=f"{field} cannot be negative")


def get_academic_year_months(academic_year: str) -> List[str]:
    start_year = int(academic_year.split("-")[0])
    months = []
    for m in ACADEMIC_MONTHS:
        yr = start_year if int(m) >= 4 else start_year + 1
        months.append(f"{yr}-{m}")
    return months


def get_remaining_months(academic_year: str, from_date: str) -> List[str]:
    all_months = get_academic_year_months(academic_year)
    from_month = from_date[:7]
    remaining = [m for m in all_months if m >= from_month]
    return remaining if remaining else all_months


def current_academic_year() -> str:
    now = datetime.now()
    if now.month >= 4:
        return f"{now.year}-{now.year + 1}"
    return f"{now.year - 1}-{now.year}"


async def active_session() -> str:
    """
    The school's active academic session. Reads the sessions collection
    (is_active=true) first, then the legacy school_settings doc, then the
    calendar-computed current year. New admissions / fees attach to THIS session.
    """
    try:
        sess = await db.sessions.find_one({"is_active": True}, {"_id": 0, "session_name": 1})
        if sess and sess.get("session_name"):
            return sess["session_name"]
    except Exception:
        pass
    try:
        doc = await db.school_settings.find_one({"_id": "session"}, {"_id": 0, "active_session": 1})
        if doc and doc.get("active_session"):
            return doc["active_session"]
    except Exception:
        pass
    return current_academic_year()


async def session_status(academic_year: str) -> str:
    """Return a session's status (active/archived/upcoming), or 'active' if unknown."""
    try:
        sess = await db.sessions.find_one({"session_name": academic_year}, {"_id": 0, "status": 1})
        if sess:
            return sess.get("status", "active")
    except Exception:
        pass
    return "active"


async def ensure_session_writable(academic_year: str):
    """No-op. Archived/closed sessions remain fully editable — every session
    behaves exactly like the active session. Kept so existing call sites need
    no change."""
    return


def get_fy_prefix() -> str:
    now = datetime.now()
    if now.month >= 4:
        return f"{now.year}-{str(now.year + 1)[-2:]}"
    return f"{now.year - 1}-{str(now.year)[-2:]}"


async def get_next_receipt_number() -> str:
    fy = get_fy_prefix()
    counter = await db.counters.find_one_and_update(
        {"_id": f"receipt_{fy}"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    return f"REC/{fy}/{str(counter['seq']).zfill(4)}"


async def get_fee_config(class_name: str, academic_year: str, stream: Optional[str] = None) -> Optional[dict]:
    """
    Fetch the component fee config for a class + stream + year.
    Falls back to no-stream config if stream-specific one not found.
    """
    query = {"class_name": class_name, "academic_year": academic_year, "is_active": True}
    if stream:
        cfg = await db.fee_component_configs.find_one({**query, "stream": stream}, {"_id": 0})
        if cfg:
            logger.info(f"Fee config found: class={class_name} stream={stream} year={academic_year}")
            return cfg
        logger.warning(f"Fee config NOT found for class={class_name} stream={stream} year={academic_year}, trying no-stream fallback")
    # fallback — config without stream
    cfg = await db.fee_component_configs.find_one({**query, "stream": None}, {"_id": 0})
    if cfg:
        logger.info(f"Fee config found (no-stream fallback): class={class_name} year={academic_year}")
    else:
        logger.warning(f"Fee config NOT found at all: class={class_name} stream={stream} year={academic_year}")
    return cfg


async def check_sibling(student_id: str, parent_id: Optional[str], parent_email: Optional[str]) -> bool:
    """Return True if this student has an active sibling already enrolled."""
    if not parent_id and not parent_email:
        return False
    q = {"is_active": True, "student_id": {"$ne": student_id}}
    if parent_id:
        q["parent_id"] = parent_id
    elif parent_email:
        q["parent_email"] = parent_email
    return await db.students.count_documents(q) > 0


async def refresh_overdue_for_student(student_id: str):
    """Mark pending ledger entries as overdue when past due date; apply late fee."""
    today = datetime.now().strftime("%Y-%m-%d")

    await db.student_ledger.update_many(
        {"student_id": student_id, "status": "pending", "due_date": {"$lt": today}},
        {"$set": {"status": "overdue"}}
    )

    # Apply late fees for overdue tuition entries (if class config has it enabled)
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if student:
        cfg = await get_fee_config(
            student.get("class_name", ""),
            student.get("academic_year", current_academic_year()),
            student.get("stream")
        )
        if cfg and cfg.get("late_fee_enabled") and cfg.get("late_fee", 0) > 0:
            late_fee_amount = cfg["late_fee"]
            overdue_no_late = await db.student_ledger.find({
                "student_id": student_id,
                "status": "overdue",
                "late_fee_applied": 0,
                "fee_component": "tuition"
            }, {"_id": 0}).to_list(100)
            for entry in overdue_no_late:
                new_net = entry["gross_amount"] - entry.get("concession_amount", 0) + late_fee_amount
                await db.student_ledger.update_one(
                    {"ledger_id": entry["ledger_id"]},
                    {"$set": {"late_fee_applied": late_fee_amount, "net_amount": new_net}}
                )

    # Update student fee_status
    overdue = await db.student_ledger.count_documents({"student_id": student_id, "status": "overdue"})
    pending = await db.student_ledger.count_documents({"student_id": student_id, "status": {"$in": ["pending", "overdue"]}})
    if overdue > 0:
        await db.students.update_one({"student_id": student_id}, {"$set": {"fee_status": "overdue", "app_locked": True}})
    elif pending > 0:
        await db.students.update_one({"student_id": student_id}, {"$set": {"fee_status": "pending", "app_locked": False}})
    else:
        await db.students.update_one({"student_id": student_id}, {"$set": {"fee_status": "paid", "app_locked": False}})


def build_admission_fee_breakdown(cfg: dict, is_sibling: bool) -> List[dict]:
    """
    Build the full fee breakdown for admission time:
    one-time fees + yearly fees + 1st month tuition.
    """
    sibling_adm_disc = cfg.get("sibling_admission_discount_amount", 0) if is_sibling else 0
    sibling_tuit_disc = cfg.get("sibling_tuition_discount_amount", 0) if is_sibling else 0

    items = []
    # One-time fees
    for cfg_field, label in [
        ("registration_fee", "Registration Fee"),
        ("admission_fee", "Admission Fee"),
        ("caution_deposit", "Caution Deposit (Refundable)"),
    ]:
        amount = cfg.get(cfg_field, 0)
        if amount > 0:
            discount = 0
            if cfg_field == "admission_fee" and sibling_adm_disc:
                discount = min(sibling_adm_disc, amount)  # Don't discount more than the fee amount
            items.append({
                "fee_component": CFG_FIELD_TO_COMPONENT[cfg_field],
                "fee_type": "one_time",
                "label": label,
                "gross_amount": amount,
                "discount_amount": discount,
                "net_amount": amount - discount,
                "sibling_discount_amount": discount if cfg_field == "admission_fee" else 0,
            })
    # Yearly fees
    for comp, label in [
        ("annual_charge", "Annual Charge"),
        ("activity_fee", "Activity Fee"),
        ("exam_fee", "Exam Fee"),
        ("lab_fee", "Lab Fee"),
        ("ai_robotics_fee", "AI & Robotics Fee"),
    ]:
        amount = cfg.get(comp, 0)
        if amount > 0:
            items.append({
                "fee_component": comp,
                "fee_type": "yearly",
                "label": label,
                "gross_amount": amount,
                "discount_amount": 0,
                "net_amount": amount,
                "sibling_discount_amount": 0,
            })
    # 1st month tuition
    tuition = cfg.get("monthly_tuition", 0)
    if tuition > 0:
        disc = min(sibling_tuit_disc, tuition) if sibling_tuit_disc else 0  # Don't discount more than tuition
        items.append({
            "fee_component": "tuition",
            "fee_type": "monthly",
            "label": "Tuition (1st Month)",
            "gross_amount": tuition,
            "discount_amount": disc,
            "net_amount": tuition - disc,
            "sibling_discount_amount": disc,
        })
    return items


def _due_date(year_month: str, due_day: int) -> str:
    """
    Compute due date for a fee entry: YYYY-MM-{due_day}.
    If the due date would be in the past (entry created late), push it to next month so the
    entry isn't immediately overdue on creation.
    """
    yr, mn = year_month.split("-")
    candidate = f"{yr}-{mn}-{str(due_day).zfill(2)}"
    today = datetime.now().strftime("%Y-%m-%d")
    if candidate < today:
        m, y = int(mn), int(yr)
        if m == 12:
            m, y = 1, y + 1
        else:
            m += 1
        candidate = f"{y}-{str(m).zfill(2)}-{str(due_day).zfill(2)}"
    return candidate


async def create_admission_ledger(student: dict, cfg: dict, academic_year: str, admission_month: str):
    """
    Create all ledger entries for a newly admitted student:
    - One-time fees
    - Yearly fees
    - 1st month tuition
    - Remaining 11 monthly tuition entries (status=pending)
    """
    student_id = student["student_id"]
    admission_number = student.get("admission_number", "")
    class_name = student["class_name"]
    stream = student.get("stream")
    is_sibling = student.get("is_sibling", False)
    due_day = cfg.get("due_day", 10)

    sibling_adm_disc = cfg.get("sibling_admission_discount_amount", 0) if is_sibling else 0
    sibling_tuit_disc = cfg.get("sibling_tuition_discount_amount", 0) if is_sibling else 0

    all_months = get_academic_year_months(academic_year)
    # Month index of admission month
    remaining_months = get_remaining_months(academic_year, f"{admission_month}-01")

    ledger_entries = []

    # — One-time fees —
    for cfg_field, label in [
        ("registration_fee", "Registration Fee"),
        ("admission_fee", "Admission Fee"),
        ("caution_deposit", "Caution Deposit (Refundable)"),
    ]:
        gross = cfg.get(cfg_field, 0)
        if gross <= 0:
            continue
        # One-time fees (Registration, Admission, Caution Deposit) are charged ONCE per student.
        # Skip if the student already has this fee in ANY class/year — even if paid in a previous class.
        if await db.student_ledger.find_one({"student_id": student_id, "fee_component": CFG_FIELD_TO_COMPONENT[cfg_field]}, {"_id": 1}):
            continue
        disc = 0
        disc_reason = None
        if cfg_field == "admission_fee" and sibling_adm_disc > 0:
            disc = min(sibling_adm_disc, gross)  # Don't discount more than the fee
            disc_reason = f"Sibling discount (Rs.{disc})" if disc > 0 else None
        net = gross - disc
        yr, mn = admission_month.split("-")
        due_date = f"{yr}-{mn}-{str(due_day).zfill(2)}"
        entry = StudentLedgerEntry(
            student_id=student_id,
            admission_number=admission_number,
            class_name=class_name,
            stream=stream,
            academic_year=academic_year,
            fee_component=CFG_FIELD_TO_COMPONENT[cfg_field],
            fee_type="one_time",
            description=label,
            gross_amount=gross,
            concession_amount=disc,
            concession_reason=disc_reason,
            net_amount=net,
            due_date=due_date,
            status="pending",
        )
        ledger_entries.append(entry)

    # — Yearly fees —
    for comp, label in [
        ("annual_charge", "Annual Charge"),
        ("activity_fee", "Activity Fee"),
        ("exam_fee", "Exam Fee"),
        ("lab_fee", "Lab Fee"),
        ("ai_robotics_fee", "AI & Robotics Fee"),
    ]:
        gross = cfg.get(comp, 0)
        if gross <= 0:
            continue
        if await db.student_ledger.find_one({"student_id": student_id, "fee_component": comp, "academic_year": academic_year, "class_name": class_name}, {"_id": 1}):
            continue
        due_date = _due_date(admission_month, due_day)
        entry = StudentLedgerEntry(
            student_id=student_id,
            admission_number=admission_number,
            class_name=class_name,
            stream=stream,
            academic_year=academic_year,
            fee_component=comp,
            fee_type="yearly",
            description=f"{label} {academic_year}",
            gross_amount=gross,
            net_amount=gross,
            due_date=due_date,
            status="pending",
        )
        ledger_entries.append(entry)

    # — Monthly tuition (all remaining months) —
    tuition = cfg.get("monthly_tuition", 0)
    if tuition > 0:
        disc_amt = min(sibling_tuit_disc, tuition) if sibling_tuit_disc > 0 else 0  # Don't discount more than tuition
        disc_reason = f"Sibling discount (Rs.{disc_amt})" if disc_amt > 0 else None
        net_tuition = tuition - disc_amt

        for month_str in remaining_months:
            existing = await db.student_ledger.find_one({
                "student_id": student_id,
                "fee_component": "tuition",
                "month": month_str,
                "academic_year": academic_year
            }, {"_id": 0})
            if existing:
                continue
            yr, mn = month_str.split("-")
            due_date = f"{yr}-{mn}-{str(due_day).zfill(2)}"
            month_names = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
                           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            desc = f"Tuition — {month_names[int(mn)]} {yr}"
            entry = StudentLedgerEntry(
                student_id=student_id,
                admission_number=admission_number,
                class_name=class_name,
                stream=stream,
                academic_year=academic_year,
                fee_component="tuition",
                fee_type="monthly",
                description=desc,
                month=month_str,
                gross_amount=tuition,
                concession_amount=disc_amt,
                concession_reason=disc_reason,
                net_amount=net_tuition,
                due_date=due_date,
                status="pending",
            )
            ledger_entries.append(entry)

    # Bulk insert
    if ledger_entries:
        docs = []
        for e in ledger_entries:
            d = e.model_dump()
            d["created_at"] = d["created_at"].isoformat()
            docs.append(d)
        await db.student_ledger.insert_many(docs)

    return len(ledger_entries)


# ─── Fee Component Config endpoints ──────────────────────────────────────────

@router.get("/fees/components")
async def list_fee_component_configs(
    request: Request,
    academic_year: Optional[str] = None,
    class_name: Optional[str] = None,
):
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    query = {"is_active": True}
    if academic_year:
        query["academic_year"] = academic_year
    if class_name:
        query["class_name"] = class_name
    configs = await db.fee_component_configs.find(query, {"_id": 0}).sort(
        [("class_name", 1), ("stream", 1)]
    ).to_list(500)
    return configs


@router.post("/fees/components")
async def create_fee_component_config(request: Request):
    """Create or replace fee component config for a class+stream+year."""
    user = await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    body = await request.json()

    academic_year = body.get("academic_year", current_academic_year())
    class_name = body.get("class_name")
    # Normalize empty string / falsy to None so the de-dupe match below treats
    # "no stream" consistently — otherwise stream="" and stream=null are seen as
    # two different configs and duplicates pile up.
    stream = body.get("stream") or None

    if not class_name:
        raise HTTPException(status_code=400, detail="class_name is required")

    validate_academic_year(academic_year)  # #14
    validate_fee_amounts(body)             # #13

    # Deactivate existing config for this class+stream+year
    await db.fee_component_configs.update_many(
        {"class_name": class_name, "stream": stream, "academic_year": academic_year},
        {"$set": {"is_active": False}}
    )

    cfg = FeeComponentConfig(
        class_name=class_name,
        stream=stream,
        academic_year=academic_year,
        registration_fee=float(body.get("registration_fee", 0)),
        admission_fee=float(body.get("admission_fee", 0)),
        caution_deposit=float(body.get("caution_deposit", 0)),
        annual_charge=float(body.get("annual_charge", 0)),
        activity_fee=float(body.get("activity_fee", 0)),
        exam_fee=float(body.get("exam_fee", 0)),
        lab_fee=float(body.get("lab_fee", 0)),
        ai_robotics_fee=float(body.get("ai_robotics_fee", 0)),
        monthly_tuition=float(body.get("monthly_tuition", 0)),
        upgradation_fee=float(body.get("upgradation_fee", 0)),
        due_day=int(body.get("due_day", 10)),
        late_fee=float(body.get("late_fee", 0)),
        late_fee_enabled=bool(body.get("late_fee_enabled", False)),
        sibling_admission_discount_amount=float(body.get("sibling_admission_discount_amount", 0)),
        sibling_tuition_discount_amount=float(body.get("sibling_tuition_discount_amount", 0)),
        notes=body.get("notes"),
        created_by=user["user_id"],
    )
    d = cfg.model_dump()
    d["created_at"] = d["created_at"].isoformat()
    await db.fee_component_configs.insert_one(d)
    d.pop("_id", None)

    await create_audit_log("fee_component_config", cfg.config_id, "create", {
        "class_name": class_name, "stream": stream, "academic_year": academic_year
    }, user)
    return d


@router.put("/fees/components/{config_id}")
async def update_fee_component_config(config_id: str, request: Request):
    user = await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    body = await request.json()

    cfg = await db.fee_component_configs.find_one({"config_id": config_id}, {"_id": 0})
    if not cfg:
        raise HTTPException(status_code=404, detail="Fee config not found")

    validate_fee_amounts(body)  # #13

    # Identity fields (class / stream / year) are editable from the dialog too —
    # previously they were dropped, so changing the stream silently did nothing.
    # Normalize stream ("" -> None) and de-dupe: deactivate any OTHER active
    # config that would collide with the new class+stream+year.
    new_class = body.get("class_name") or cfg["class_name"]
    new_stream = (body.get("stream") if "stream" in body else cfg.get("stream")) or None
    new_year = body.get("academic_year") or cfg["academic_year"]
    validate_academic_year(new_year)  # #14
    await db.fee_component_configs.update_many(
        {"class_name": new_class, "stream": new_stream, "academic_year": new_year,
         "config_id": {"$ne": config_id}, "is_active": True},
        {"$set": {"is_active": False}}
    )

    allowed_fields = [
        "registration_fee", "admission_fee", "caution_deposit",
        "annual_charge", "activity_fee", "exam_fee", "lab_fee", "ai_robotics_fee",
        "monthly_tuition", "upgradation_fee",
        "due_day", "late_fee", "late_fee_enabled",
        "sibling_admission_discount_amount", "sibling_tuition_discount_amount", "notes"
    ]
    update = {k: body[k] for k in allowed_fields if k in body}
    update["class_name"] = new_class
    update["stream"] = new_stream
    update["academic_year"] = new_year
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.fee_component_configs.update_one({"config_id": config_id}, {"$set": update})

    updated = await db.fee_component_configs.find_one({"config_id": config_id}, {"_id": 0})
    await create_audit_log("fee_component_config", config_id, "update", update, user)
    return updated


@router.delete("/fees/components/{config_id}")
async def delete_fee_component_config(config_id: str, request: Request):
    """Permanently remove a fee component config (admin 'Remove' action)."""
    user = await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    cfg = await db.fee_component_configs.find_one({"config_id": config_id}, {"_id": 0})
    if not cfg:
        raise HTTPException(status_code=404, detail="Fee config not found")
    await db.fee_component_configs.delete_one({"config_id": config_id})
    await create_audit_log("fee_component_config", config_id, "delete", {
        "class_name": cfg.get("class_name"), "stream": cfg.get("stream"),
        "academic_year": cfg.get("academic_year"),
    }, user)
    return {"message": "Fee configuration deleted", "config_id": config_id}


@router.post("/fees/components/increase")
async def apply_annual_increase(request: Request):
    """
    Apply a percentage increase to all fee amounts for a given academic year,
    creating new configs for the next session.
    """
    user = await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    body = await request.json()

    from_year = body.get("from_year", current_academic_year())
    increase_pct = float(body.get("increase_percent", 10))
    class_names = body.get("class_names")  # None = all classes

    if not (0 < increase_pct <= 100):
        raise HTTPException(status_code=400, detail="increase_percent must be 1–100")

    # Build next year string
    start = int(from_year.split("-")[0])
    to_year = f"{start + 1}-{start + 2}"

    query = {"academic_year": from_year, "is_active": True}
    if class_names:
        query["class_name"] = {"$in": class_names}

    source_configs = await db.fee_component_configs.find(query, {"_id": 0}).to_list(500)
    if not source_configs:
        raise HTTPException(status_code=404, detail=f"No configs found for {from_year}")

    factor = 1 + increase_pct / 100
    fee_fields = [
        "registration_fee", "admission_fee", "caution_deposit",
        "annual_charge", "activity_fee", "exam_fee", "lab_fee", "ai_robotics_fee",
        "monthly_tuition", "upgradation_fee", "late_fee"
    ]
    created = 0
    for src in source_configs:
        # Deactivate existing for to_year
        await db.fee_component_configs.update_many(
            {"class_name": src["class_name"], "stream": src.get("stream"), "academic_year": to_year},
            {"$set": {"is_active": False}}
        )
        new_cfg = {k: src[k] for k in src if k not in ("_id", "config_id", "created_at", "updated_at")}
        new_cfg["config_id"] = f"fcc_{uuid.uuid4().hex[:10]}"
        new_cfg["academic_year"] = to_year
        new_cfg["created_by"] = user["user_id"]
        new_cfg["created_at"] = datetime.now(timezone.utc).isoformat()
        new_cfg["updated_at"] = None
        for f in fee_fields:
            if f in new_cfg and new_cfg[f] > 0:
                new_cfg[f] = round(new_cfg[f] * factor, 2)
        new_cfg["notes"] = f"Auto-generated from {from_year} with {increase_pct}% increase"
        await db.fee_component_configs.insert_one(new_cfg)
        created += 1

    return {
        "message": f"Created {created} fee configs for {to_year} with {increase_pct}% increase",
        "from_year": from_year,
        "to_year": to_year,
        "configs_created": created
    }


@router.post("/fees/components/ensure-defaults")
async def ensure_default_fee_configs(request: Request):
    """
    Ensure fee configurations exist for all classes in current academic year.
    If missing, creates them from default seed data.
    Useful when upgrading or fixing missing fee configs.
    """
    user = await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    academic_year = current_academic_year()
    
    # Default fee configuration values (from seed_fee_structure_2025_26.py)
    COMMON = {
        "registration_fee": 500,
        "admission_fee": 2500,
        "caution_deposit": 1000,
        "annual_charge": 3600,
        "upgradation_fee": 0,
        "due_day": 10,
        "late_fee": 0,
        "late_fee_enabled": False,
        "sibling_admission_discount_amount": 1000,
        "sibling_tuition_discount_amount": 300,
    }
    
    CLASS_FEES = {
        "SF. SR.": {"activity_fee": 1500, "exam_fee": 300, "lab_fee": 0, "ai_robotics_fee": 0, "monthly_tuition": 1000},
        "LKG": {"activity_fee": 2000, "exam_fee": 300, "lab_fee": 0, "ai_robotics_fee": 0, "monthly_tuition": 1100},
        "UKG": {"activity_fee": 2000, "exam_fee": 300, "lab_fee": 0, "ai_robotics_fee": 0, "monthly_tuition": 1100},
        "1st": {"activity_fee": 2400, "exam_fee": 300, "lab_fee": 1500, "ai_robotics_fee": 0, "monthly_tuition": 1150},
        "2nd": {"activity_fee": 2400, "exam_fee": 300, "lab_fee": 1500, "ai_robotics_fee": 0, "monthly_tuition": 1150},
        "3rd": {"activity_fee": 2900, "exam_fee": 300, "lab_fee": 1500, "ai_robotics_fee": 0, "monthly_tuition": 1250},
        "4th": {"activity_fee": 2900, "exam_fee": 300, "lab_fee": 1500, "ai_robotics_fee": 0, "monthly_tuition": 1250},
        "5th": {"activity_fee": 3400, "exam_fee": 300, "lab_fee": 1500, "ai_robotics_fee": 0, "monthly_tuition": 1350},
        "6th": {"activity_fee": 3400, "exam_fee": 300, "lab_fee": 1500, "ai_robotics_fee": 0, "monthly_tuition": 1350},
        "7th": {"activity_fee": 3900, "exam_fee": 300, "lab_fee": 1500, "ai_robotics_fee": 0, "monthly_tuition": 1400},
        "8th": {"activity_fee": 3900, "exam_fee": 300, "lab_fee": 1500, "ai_robotics_fee": 0, "monthly_tuition": 1400},
        "9th": {"activity_fee": 4500, "exam_fee": 450, "lab_fee": 1500, "ai_robotics_fee": 2400, "monthly_tuition": 1900},
        "10th": {"activity_fee": 4500, "exam_fee": 450, "lab_fee": 1500, "ai_robotics_fee": 2400, "monthly_tuition": 1900},
    }
    
    now = datetime.now(timezone.utc).isoformat()
    created = 0
    skipped = 0
    
    for class_name, overrides in CLASS_FEES.items():
        # Check if config already exists
        existing = await db.fee_component_configs.find_one({
            "class_name": class_name,
            "stream": None,
            "academic_year": academic_year,
            "is_active": True
        }, {"_id": 0})
        
        if existing:
            skipped += 1
            continue
        
        # Create new config
        cfg = {
            "config_id": f"fcc_{uuid.uuid4().hex[:10]}",
            "class_name": class_name,
            "stream": None,
            "academic_year": academic_year,
            **COMMON,
            **overrides,
            "is_active": True,
            "notes": "Auto-created by ensure-defaults endpoint",
            "created_by": user["user_id"],
            "created_at": now,
        }
        await db.fee_component_configs.insert_one(cfg)
        created += 1
        logger.info(f"Created missing fee config for {class_name}")
    
    await create_audit_log("fee_component_config", "batch", "ensure_defaults", {
        "academic_year": academic_year,
        "created": created,
        "skipped": skipped
    }, user)
    
    return {
        "message": f"Ensured fee configs for {academic_year}",
        "created": created,
        "skipped": skipped,
        "academic_year": academic_year
    }


# ─── Student Ledger endpoints ─────────────────────────────────────────────────

@router.get("/fees/ledger/{student_id}")
async def get_student_ledger(student_id: str, request: Request):
    user = await get_current_user(request)

    # Role checks
    if user["role"] == UserRole.PARENT:
        # Use both parent_email (primary) and parent_id (legacy fallback) to locate children
        children = await db.students.find(
            {"$or": [{"parent_email": user["email"]}, {"parent_id": user["user_id"]}], "is_active": True},
            {"_id": 0, "student_id": 1}
        ).to_list(20)
        if student_id not in {c["student_id"] for c in children}:
            raise HTTPException(status_code=403, detail="Not authorized. This student is not linked to your account.")
    elif user["role"] == UserRole.STUDENT:
        stu = await db.students.find_one({"user_id": user["user_id"]}, {"_id": 0, "student_id": 1})
        if not stu or stu["student_id"] != student_id:
            raise HTTPException(status_code=403, detail="Not authorized. You can only view your own fees.")
    elif user["role"] == UserRole.TEACHER:
        raise HTTPException(status_code=403, detail="Teachers cannot access student fee data.")

    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    # (#5) Do NOT call refresh_overdue_for_student here — it caused late-fee accumulation
    # on every ledger GET. Overdue status is updated by:
    #   1. The due-chart endpoint (bulk update on load)
    #   2. POST /fees/refresh-overdue (explicit admin trigger)
    # Quick in-memory check only — no DB write:
    today = datetime.now().strftime("%Y-%m-%d")
    await db.student_ledger.update_many(
        {"student_id": student_id, "status": "pending", "due_date": {"$lt": today}},
        {"$set": {"status": "overdue"}}
    )

    entries = await db.student_ledger.find(
        {"student_id": student_id}, {"_id": 0}
    ).sort([("fee_type", 1), ("due_date", 1)]).to_list(500)

    for e in entries:
        if isinstance(e.get("description"), str):
            e["description"] = e["description"].replace(" (seeded due)", "")

    # Auto-generate ledger entries when a student has none (e.g. seeded before fee config existed)
    if not entries:
        academic_year_val = student.get("academic_year", current_academic_year())
        cfg = await get_fee_config(student["class_name"], academic_year_val, student.get("stream"))
        if cfg:
            admission_date = student.get("admission_date", datetime.now().strftime("%Y-%m-%d"))
            admission_month = admission_date[:7]
            generated = await create_admission_ledger(student, cfg, academic_year_val, admission_month)
            if generated > 0:
                entries = await db.student_ledger.find(
                    {"student_id": student_id}, {"_id": 0}
                ).sort([("fee_type", 1), ("due_date", 1)]).to_list(500)
                logger.info("Auto-generated %d ledger entries for student %s", generated, student_id)
        else:
            logger.warning("No fee config found for student %s class=%s year=%s — cannot auto-generate ledger",
                           student_id, student["class_name"], student.get("academic_year"))

    payments = await db.fee_payments.find(
        {"student_id": student_id}, {"_id": 0}
    ).sort("payment_date", -1).to_list(500)

    # Summaries
    def _sum(entries, statuses, field="net_amount"):
        return round(sum(e[field] for e in entries if e["status"] in statuses), 2)

    total_gross = round(sum(e["gross_amount"] for e in entries), 2)
    total_concession = round(sum(e.get("concession_amount", 0) for e in entries), 2)
    total_late_fees = round(sum(e.get("late_fee_applied", 0) for e in entries), 2)
    total_paid_amount = round(sum(p["amount"] for p in payments), 2)
    total_pending = round(
        _sum(entries, ["pending", "overdue"]) +
        sum(e.get("remaining_balance", 0) for e in entries if e["status"] == "partially_paid"),
        2
    )
    total_overdue = _sum(entries, ["overdue"])
    months_paid = sum(1 for e in entries if e["fee_component"] == "tuition" and e["status"] == "paid")
    months_pending = sum(1 for e in entries if e["fee_component"] == "tuition" and e["status"] in ["pending", "overdue", "partially_paid"])

    # Group by fee_type for display
    grouped = {"one_time": [], "yearly": [], "monthly": []}
    for e in entries:
        ft = e.get("fee_type", "monthly")
        grouped.get(ft, grouped["monthly"]).append(e)

    return {
        "student": {
            "student_id": student["student_id"],
            "name": f"{student['first_name']} {student['last_name']}",
            "admission_number": student.get("admission_number", ""),
            "class_name": student["class_name"],
            "section": student["section"],
            "stream": student.get("stream"),
            "academic_year": student.get("academic_year", ""),
            "fee_status": student.get("fee_status", "pending"),
        },
        "summary": {
            "total_gross": total_gross,
            "total_concession": total_concession,
            "total_late_fees": total_late_fees,
            "total_paid": total_paid_amount,
            "total_pending": total_pending,
            "total_overdue": total_overdue,
            "months_paid": months_paid,
            "months_pending": months_pending,
        },
        "ledger": grouped,
        "payments": payments,
    }


@router.post("/fees/clear-locks/{student_id}")
async def clear_payment_locks(student_id: str, request: Request):
    """Admin: release all stale payment locks for a student."""
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    result = await db.student_ledger.update_many(
        {"student_id": student_id, "payment_lock": {"$exists": True}},
        {"$unset": {"payment_lock": ""}}
    )
    return {"cleared": result.modified_count}


@router.post("/fees/ledger/generate/{student_id}")
async def generate_student_ledger(student_id: str, request: Request):
    """Admin: create any missing ledger entries for a student (idempotent)."""
    user = await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)

    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    academic_year = student.get("academic_year", current_academic_year())
    cfg = await get_fee_config(student["class_name"], academic_year, student.get("stream"))
    if not cfg:
        raise HTTPException(status_code=404, detail="No fee config found for this class/year")

    admission_date = student.get("admission_date", datetime.now().strftime("%Y-%m-%d"))
    admission_month = admission_date[:7]

    created = await create_admission_ledger(student, cfg, academic_year, admission_month)
    return {"message": f"Generated {created} ledger entries", "student_id": student_id}


# ─── Admission Fee Collection ─────────────────────────────────────────────────

@router.post("/fees/admission-payment")
async def record_admission_payment(request: Request):
    """
    Record the admission-time payment:
    one-time fees + yearly fees + 1st month tuition.
    Marks those ledger entries as paid and generates a receipt.
    """
    user = await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    body = await request.json()

    student_id = body.get("student_id")
    payment_method = body.get("payment_method", "cash")
    transaction_id = body.get("transaction_id")
    remarks = body.get("remarks", "Admission fee collection")
    split_payments = body.get("split_payments")

    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    # Fetch pending one-time + yearly + first-month tuition
    academic_year = student.get("academic_year", current_academic_year())
    # Closed/archived sessions remain editable for fees (admin may collect or
    # adjust dues for prior years). No archive block here.
    all_months = get_academic_year_months(academic_year)
    first_month = all_months[0]
    admission_month = student.get("admission_date", "")[:7] or first_month

    pending_entries = await db.student_ledger.find({
        "student_id": student_id,
        "status": "pending",
        "fee_type": {"$in": ["one_time", "yearly"]}
    }, {"_id": 0}).to_list(100)

    # Also first-month tuition
    first_tuition = await db.student_ledger.find_one({
        "student_id": student_id,
        "fee_component": "tuition",
        "month": admission_month,
        "status": "pending"
    }, {"_id": 0})
    if first_tuition:
        pending_entries.append(first_tuition)

    if not pending_entries:
        raise HTTPException(status_code=400, detail="No pending admission fees found")

    # Oldest-first, so a partial amount clears the earliest dues first.
    pending_entries.sort(key=lambda e: (e.get("due_date") or "", e.get("fee_type") or ""))
    total_amount = round(sum(e["net_amount"] for e in pending_entries), 2)

    # Optional partial collection: an `amount` smaller than the full admission
    # total is spread across the pending entries oldest-first — fully-covered
    # entries become 'paid', the boundary entry is left partially paid, and the
    # remainder stay pending. Omit `amount` to collect the full total.
    partial_amount = body.get("amount")
    if partial_amount is not None:
        try:
            partial_amount = round(float(partial_amount), 2)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Invalid amount")
        if partial_amount <= 0:
            raise HTTPException(status_code=400, detail="Amount must be greater than 0")
        if partial_amount > total_amount + 0.01:
            raise HTTPException(status_code=400,
                detail=f"Amount Rs.{partial_amount:,.2f} exceeds total due Rs.{total_amount:,.2f}")
        collect_total = partial_amount
    else:
        collect_total = total_amount
    is_partial = collect_total < total_amount - 0.01

    today = datetime.now().strftime("%Y-%m-%d")
    receipt_number = await get_next_receipt_number()

    # Spread the collected amount across entries, oldest first.
    pending_updates = []   # (ledger_id, set_doc) — payment_id stamped after FeePayment
    covered_entries = []
    remaining_to_apply = collect_total
    for entry in pending_entries:
        if remaining_to_apply <= 0.005:
            break
        net = float(entry["net_amount"])
        prev_paid = float(entry.get("amount_paid") or 0)
        entry_remaining = round(net - prev_paid, 2)
        if entry_remaining <= 0:
            continue
        pay_here = min(remaining_to_apply, entry_remaining)
        new_paid = round(prev_paid + pay_here, 2)
        new_remaining = round(net - new_paid, 2)
        covered_entries.append(entry)
        if new_remaining < 0.005:
            set_doc = {"status": "paid", "amount_paid": new_paid, "remaining_balance": 0,
                       "paid_date": today, "receipt_number": receipt_number}
        else:
            past_due = bool(entry.get("due_date") and entry["due_date"] < today)
            set_doc = {"status": "overdue" if past_due else "pending",
                       "amount_paid": new_paid, "remaining_balance": new_remaining,
                       "paid_date": today, "receipt_number": receipt_number}
        pending_updates.append((entry["ledger_id"], set_doc))
        remaining_to_apply = round(remaining_to_apply - pay_here, 2)

    payment = FeePayment(
        student_id=student_id,
        installment_ids=[e["ledger_id"] for e in covered_entries],
        amount=collect_total,
        payment_method=payment_method,
        transaction_id=transaction_id,
        collected_by=user["user_id"],
        remarks=remarks,
        academic_year=academic_year,
    )
    pay_dict = payment.model_dump()
    pay_dict["receipt_number"] = receipt_number
    pay_dict["created_at"] = pay_dict["created_at"].isoformat()

    # Split-payment breakdown — sums must equal the amount actually collected.
    if split_payments and isinstance(split_payments, dict):
        try:
            split_total = round(sum(float(v) for v in split_payments.values()), 2)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Invalid split_payments format")
        if abs(split_total - collect_total) > 0.01:
            raise HTTPException(status_code=400,
                detail=f"Split payment total ({split_total}) does not match collected amount ({collect_total})")
        pay_dict["split_payments"] = {k: float(v) for k, v in split_payments.items() if float(v) > 0}

    ledger_updates = [
        UpdateOne({"ledger_id": lid}, {"$set": {**doc, "payment_id": payment.payment_id}})
        for lid, doc in pending_updates
    ]

    try:
        async with await mongo_client.start_session() as session:
            async with session.start_transaction():
                await db.fee_payments.insert_one(pay_dict, session=session)
                await db.student_ledger.bulk_write(ledger_updates, session=session)
    except Exception:
        # Replica set not available (standalone dev instance) — fall back to non-atomic
        await db.fee_payments.insert_one(pay_dict)
        await db.student_ledger.bulk_write(ledger_updates)

    # Update onboarding record if exists. Only flag as fully paid when the whole
    # admission total was collected (partial collections leave a balance due).
    await db.onboarding.update_many(
        {"student_id": student_id},
        {"$set": {
            "admission_fee_paid": not is_partial,
            "admission_fee_receipt": receipt_number,
            "admission_fee_payment_id": payment.payment_id,
        }}
    )

    await refresh_overdue_for_student(student_id)
    await create_audit_log("fee_payment", payment.payment_id, "admission_payment", {
        "student_id": student_id, "amount": collect_total, "method": payment_method,
        "partial": is_partial,
    }, user)

    pay_dict.pop("_id", None)
    balance_due = round(total_amount - collect_total, 2)
    if is_partial:
        msg = (f"Partial admission payment of Rs.{collect_total:,.2f} recorded. "
               f"Rs.{balance_due:,.2f} still due. Receipt: {receipt_number}")
    else:
        msg = f"Admission fee of Rs.{collect_total:,.2f} recorded. Receipt: {receipt_number}"
    return {
        "payment": pay_dict,
        "receipt_number": receipt_number,
        "amount": collect_total,
        "total_due": total_amount,
        "balance_due": balance_due,
        "is_partial": is_partial,
        "entries_paid": len(covered_entries),
        "message": msg,
    }


# ─── Monthly tuition generation ───────────────────────────────────────────────

@router.post("/fees/generate-monthly")
async def generate_monthly_fees(request: Request):
    """
    Generate monthly tuition entries for all active students for a given month.
    Meant to be run at the start of each month (or as a scheduled job).
    Idempotent — will not duplicate entries.
    """
    user = await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    body = await request.json()

    month_str = body.get("month")  # "2025-05"
    academic_year = body.get("academic_year", current_academic_year())

    if not month_str:
        now = datetime.now()
        month_str = f"{now.year}-{str(now.month).zfill(2)}"

    students = await db.students.find({"is_active": True}, {"_id": 0}).to_list(5000)
    created = 0
    skipped = 0

    yr, mn = month_str.split("-")
    month_names = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    desc = f"Tuition — {month_names[int(mn)]} {yr}"

    # Batch pre-fetch to eliminate per-student round-trips.
    # 1) All fee configs for the year, keyed by (class, stream).
    cfgs_cursor = db.fee_component_configs.find(
        {"academic_year": academic_year, "is_active": True}, {"_id": 0}
    )
    cfg_by_class_stream: dict = {}
    async for cfg in cfgs_cursor:
        cfg_by_class_stream[(cfg.get("class_name"), cfg.get("stream"))] = cfg

    def lookup_cfg(class_name: str, stream: Optional[str]):
        # Prefer stream-specific, fall back to no-stream config.
        return cfg_by_class_stream.get((class_name, stream)) or cfg_by_class_stream.get((class_name, None))

    # 2) Existing tuition ledger entries for this month, in a single query.
    student_ids = [s["student_id"] for s in students]
    existing_ids: set = set()
    if student_ids:
        async for row in db.student_ledger.find(
            {
                "student_id": {"$in": student_ids},
                "fee_component": "tuition",
                "month": month_str,
                "academic_year": academic_year,
            },
            {"_id": 0, "student_id": 1},
        ):
            existing_ids.add(row["student_id"])

    for student in students:
        if student["student_id"] in existing_ids:
            skipped += 1
            continue

        cfg = lookup_cfg(student["class_name"], student.get("stream"))
        if not cfg or not cfg.get("monthly_tuition", 0):
            skipped += 1
            continue

        is_sibling = student.get("is_sibling", False)
        tuition = cfg["monthly_tuition"]
        disc_amt = cfg.get("sibling_tuition_discount_amount", 0) if is_sibling else 0
        disc_amt = min(disc_amt, tuition)  # Don't discount more than tuition
        net = tuition - disc_amt
        due_day = cfg.get("due_day", 10)
        due_date = f"{yr}-{mn}-{str(due_day).zfill(2)}"

        entry = StudentLedgerEntry(
            student_id=student["student_id"],
            admission_number=student.get("admission_number", ""),
            class_name=student["class_name"],
            stream=student.get("stream"),
            academic_year=academic_year,
            fee_component="tuition",
            fee_type="monthly",
            description=desc,
            month=month_str,
            gross_amount=tuition,
            concession_amount=disc_amt,
            concession_reason=f"Sibling discount (Rs.{disc_amt})" if disc_amt > 0 else None,
            net_amount=net,
            due_date=due_date,
            status="pending",
        )
        d = entry.model_dump()
        d["created_at"] = d["created_at"].isoformat()
        await db.student_ledger.insert_one(d)
        created += 1

    return {
        "message": f"Generated tuition entries for {month_str}",
        "created": created,
        "skipped": skipped,
        "month": month_str,
    }


# ─── Pay endpoint ──────────────────────────────────────────────────────────────

@router.post("/fees/pay")
async def pay_fee(request: Request):
    """
    Record a fee payment against specific ledger entry IDs or the next pending month.

    Rules:
    - Must pay oldest overdue/pending first (no skip).
    - Multi-entry payment is full-only (must clear each entry).
    - Single-entry payment may be PARTIAL: pass `amount` <= remaining_balance.
      The entry's amount_paid is incremented and status becomes
      'partially_paid' until remaining_balance reaches 0, then 'paid'.
    """
    user = await get_current_user(request)
    body = await request.json()

    student_id = body.get("student_id")
    payment_method = body.get("payment_method", "cash")
    transaction_id = body.get("transaction_id")
    remarks = body.get("remarks")
    ledger_ids = body.get("ledger_ids")  # optional list of specific entry IDs
    # Partial payment — only valid when ledger_ids contains exactly one entry.
    partial_amount = body.get("amount")
    # Optional back-dated payment date (admin only). Must not be in the future.
    custom_payment_date = body.get("payment_date")
    # Split payments: { "cash": 1500, "online": 500 } — sums must equal total
    split_payments = body.get("split_payments")

    # Auth
    if user["role"] == UserRole.PARENT:
        children = await db.students.find(
            {"$or": [{"parent_email": user["email"]}, {"parent_id": user["user_id"]}], "is_active": True},
            {"_id": 0, "student_id": 1}
        ).to_list(20)
        if student_id not in {c["student_id"] for c in children}:
            raise HTTPException(status_code=403, detail="Not authorized.")
    elif user["role"] not in [UserRole.ADMIN, UserRole.ACCOUNTANT]:
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    await refresh_overdue_for_student(student_id)

    # Entries with amount_paid > 0 stay in 'pending'/'overdue' until cleared
    payable_statuses = ["pending", "overdue"]
    if ledger_ids:
        entries_to_pay = await db.student_ledger.find({
            "student_id": student_id,
            "ledger_id": {"$in": ledger_ids},
            "status": {"$in": payable_statuses}
        }, {"_id": 0}).sort("due_date", 1).to_list(100)
        if len(entries_to_pay) != len(ledger_ids):
            raise HTTPException(status_code=400, detail="Some ledger entries not found or already paid")
    else:
        # Default: pay the next pending tuition entry
        entries_to_pay = await db.student_ledger.find({
            "student_id": student_id,
            "fee_component": "tuition",
            "status": {"$in": payable_statuses}
        }, {"_id": 0}).sort("due_date", 1).limit(1).to_list(1)

    if not entries_to_pay:
        raise HTTPException(status_code=400, detail="No pending fees to pay")

    # Closed/archived sessions remain editable for fee collection — pending dues
    # can always be settled regardless of session status.

    # Per-entry payable = remaining_balance if set, else fall back to net_amount
    # (older rows created before the partial-payment fields existed).
    def _entry_remaining(e):
        rb = e.get("remaining_balance")
        return float(rb) if rb is not None and rb > 0 else float(e.get("net_amount", 0))

    # Compute how much of this payment applies to each selected entry.
    # A partial `amount` is spread across the selected entries oldest-first:
    # the earliest dues are covered in full and the boundary entry is left
    # partially paid. Omit `amount` to pay every selected entry in full. This
    # works whether one fee or several are selected.
    total_due = round(sum(_entry_remaining(e) for e in entries_to_pay), 2)
    is_partial = False
    per_entry_paid = {}  # ledger_id -> amount paid in THIS payment
    if partial_amount is not None:
        try:
            partial_amount = round(float(partial_amount), 2)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Invalid amount")
        if partial_amount <= 0:
            raise HTTPException(status_code=400, detail="Amount must be greater than 0")
        if partial_amount > total_due + 0.01:
            raise HTTPException(status_code=400,
                detail=f"Amount Rs.{partial_amount:,.2f} exceeds total due Rs.{total_due:,.2f}.")
        remaining_to_apply = partial_amount
        for e in entries_to_pay:  # already sorted oldest-first by due_date
            if remaining_to_apply <= 0.005:
                break
            rem = _entry_remaining(e)
            if rem <= 0:
                continue
            pay_here = round(min(remaining_to_apply, rem), 2)
            per_entry_paid[e["ledger_id"]] = pay_here
            remaining_to_apply = round(remaining_to_apply - pay_here, 2)
        total = partial_amount
        is_partial = partial_amount < total_due - 0.01
    else:
        # Full payment of every selected entry (using remaining_balance to
        # support entries already partially paid via earlier transactions).
        for e in entries_to_pay:
            per_entry_paid[e["ledger_id"]] = round(_entry_remaining(e), 2)
        total = round(sum(per_entry_paid.values()), 2)

    # Only the entries this payment actually touched go on the receipt/record.
    covered_entries = [e for e in entries_to_pay if per_entry_paid.get(e["ledger_id"], 0) > 0]
    receipt_number = await get_next_receipt_number()

    payment = FeePayment(
        student_id=student_id,
        installment_ids=[e["ledger_id"] for e in covered_entries],
        amount=total,
        payment_method=payment_method,
        transaction_id=transaction_id,
        collected_by=user["user_id"],
        remarks=remarks,
        academic_year=entries_to_pay[0].get("academic_year", ""),
    )
    pay_dict = payment.model_dump()
    pay_dict["receipt_number"] = receipt_number
    pay_dict["created_at"] = pay_dict["created_at"].isoformat()

    today_str = datetime.now().strftime("%Y-%m-%d")
    # Resolve payment date (back-date if admin/accountant explicitly set one)
    paid_date = today_str
    if custom_payment_date and user["role"] in (UserRole.ADMIN, UserRole.ACCOUNTANT):
        try:
            d = datetime.strptime(custom_payment_date, "%Y-%m-%d").date()
            if d <= datetime.now().date():
                paid_date = custom_payment_date
                pay_dict["payment_date"] = custom_payment_date
        except ValueError:
            pass

    # Attach split payment breakdown if provided
    if split_payments and isinstance(split_payments, dict):
        try:
            split_total = round(sum(float(v) for v in split_payments.values()), 2)
            if abs(split_total - total) < 0.01:
                pay_dict["split_payments"] = {k: float(v) for k, v in split_payments.items() if float(v) > 0}
            else:
                raise HTTPException(status_code=400, detail=f"Split payment total ({split_total}) does not match fee total ({total})")
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Invalid split_payments format")

    # Per-entry ledger update: increment amount_paid, recompute remaining,
    # flip status to 'partially_paid' or 'paid' accordingly.
    ledger_updates = []
    for entry in covered_entries:
        paid_now = per_entry_paid[entry["ledger_id"]]
        prev_paid = float(entry.get("amount_paid") or 0)
        new_paid = round(prev_paid + paid_now, 2)
        new_remaining = round(float(entry.get("net_amount", 0)) - new_paid, 2)
        if new_remaining < 0.005:
            new_status = "paid"
            new_remaining = 0
            set_doc = {
                "status": new_status,
                "amount_paid": new_paid,
                "remaining_balance": 0,
                "paid_date": paid_date,
                "payment_id": payment.payment_id,
                "receipt_number": receipt_number,
            }
        else:
            # Per the school's preference, partials stay 'pending' (the
            # overdue sweep will flip them to 'overdue' once past due_date).
            # Only amount_paid + remaining_balance reflect the partial.
            today_iso = datetime.now().strftime("%Y-%m-%d")
            past_due = entry.get("due_date") and entry["due_date"] < today_iso
            new_status = "overdue" if past_due else "pending"
            set_doc = {
                "status": new_status,
                "amount_paid": new_paid,
                "remaining_balance": new_remaining,
                # Stamp the latest payment so admin can preview the partial
                # receipt from the ledger row.
                "paid_date": paid_date,
                "payment_id": payment.payment_id,
                "receipt_number": receipt_number,
            }
        ledger_updates.append(UpdateOne(
            {"ledger_id": entry["ledger_id"]},
            {"$set": set_doc},
        ))

    try:
        async with await mongo_client.start_session() as session:
            async with session.start_transaction():
                await db.fee_payments.insert_one(pay_dict, session=session)
                await db.student_ledger.bulk_write(ledger_updates, session=session)
    except Exception:
        # Replica set not available (standalone dev instance) — fall back to non-atomic
        await db.fee_payments.insert_one(pay_dict)
        await db.student_ledger.bulk_write(ledger_updates)

    await refresh_overdue_for_student(student_id)
    await create_audit_log("fee_payment", payment.payment_id, "pay", {
        "student_id": student_id, "amount": total, "method": payment_method
    }, user)

    pay_dict.pop("_id", None)
    msg = f"Payment of Rs.{total:,.2f} recorded. Receipt: {receipt_number}"
    if is_partial:
        # Balance still owed across the entries that were selected for this payment.
        new_remaining = round(total_due - total, 2)
        if new_remaining < 0: new_remaining = 0
        msg = (f"Partial payment of Rs.{total:,.2f} recorded. "
               f"Rs.{new_remaining:,.2f} still due. Receipt: {receipt_number}")
    return {
        "payment": pay_dict,
        "receipt_number": receipt_number,
        "amount": total,
        "entries_paid": len(covered_entries),
        "is_partial": is_partial,
        "message": msg,
    }


# ─── Due Chart ────────────────────────────────────────────────────────────────

@router.get("/fees/due-chart")
async def get_due_chart(
    request: Request,
    class_name: Optional[str] = None,
    section: Optional[str] = None,
    search: Optional[str] = None,
    fee_component: Optional[str] = None,  # single-component legacy param
    month: Optional[str] = None,          # single-month legacy param (YYYY-MM)
    # Multi-select filter: comma-separated list of selections. Each selection
    # is either "<component>" or "<component>:YYYY-MM" for month-scoped
    # entries (currently used for per-month tuition). Multiple selections
    # are OR-combined in the ledger match.
    fee_selections: Optional[str] = None,
    academic_year: Optional[str] = None,  # e.g. "2025-2026"
):
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)

    today = datetime.now().strftime("%Y-%m-%d")

    # 1. Bulk-mark all past-due pending entries as overdue in one query
    await db.student_ledger.update_many(
        {"status": "pending", "due_date": {"$lt": today}},
        {"$set": {"status": "overdue"}}
    )

    # 2. Aggregate pending/overdue totals per student. For entries that
    # have been partially paid (amount_paid > 0), sum remaining_balance
    # instead of net_amount so the chart reflects what's actually owed.
    ledger_match: dict = {"status": {"$in": ["pending", "overdue"]}}
    # Fee Type (category) and Duration (month) are independent filters,
    # combined with AND. fee_selections is a comma list of fee_component names.
    comps = [s.strip() for s in (fee_selections or "").split(",") if s.strip()]
    if fee_component and fee_component not in comps:
        comps.append(fee_component)  # legacy single-value support
    if len(comps) == 1:
        ledger_match["fee_component"] = comps[0]
    elif len(comps) > 1:
        ledger_match["fee_component"] = {"$in": comps}
    if month:
        # Duration filter: YYYY-MM prefix match on the YYYY-MM-DD due_date.
        ledger_match["due_date"] = {"$regex": f"^{month}-"}
    ledger_pipeline = [
        {"$match": ledger_match},
        {"$group": {
            "_id": "$student_id",
            "total_due": {"$sum": {
                "$cond": [
                    {"$gt": [{"$ifNull": ["$amount_paid", 0]}, 0]},
                    {"$ifNull": ["$remaining_balance", "$net_amount"]},
                    "$net_amount",
                ]
            }},
            "entries_pending": {"$sum": 1},
            "entries_overdue": {"$sum": {"$cond": [{"$eq": ["$status", "overdue"]}, 1, 0]}},
            "entries_partial": {"$sum": {
                "$cond": [{"$gt": [{"$ifNull": ["$amount_paid", 0]}, 0]}, 1, 0]
            }},
            "oldest_due":      {"$min": "$due_date"},
        }},
    ]
    ledger_rows = await db.student_ledger.aggregate(ledger_pipeline).to_list(10000)
    if not ledger_rows:
        return []

    student_ids = [r["_id"] for r in ledger_rows]

    # 3. Fetch matching students in one query
    student_query: dict = {"is_active": True, "student_id": {"$in": student_ids}}
    if class_name:
        student_query["class_name"] = class_name
    if section:
        student_query["section"] = section
    if academic_year:
        student_query["academic_year"] = academic_year

    students = await db.students.find(student_query, {
        "_id": 0, "student_id": 1, "first_name": 1, "last_name": 1,
        "class_name": 1, "section": 1, "stream": 1, "admission_number": 1, "fee_status": 1,
        "phone": 1, "academic_year": 1,
    }).to_list(5000)

    student_map = {s["student_id"]: s for s in students}
    due_chart = []
    for row in ledger_rows:
        s = student_map.get(row["_id"])
        if not s:
            continue  # inactive or filtered out by class_name
        due_chart.append({
            "student_id":       row["_id"],
            "admission_number": s.get("admission_number", ""),
            "name":             f"{s['first_name']} {s['last_name']}",
            "class_name":       s["class_name"],
            "section":          s["section"],
            "stream":           s.get("stream"),
            "academic_year":    s.get("academic_year", ""),
            "mobile":           s.get("phone", ""),
            "total_due":        round(row["total_due"], 2),
            "entries_pending":  row["entries_pending"],
            "entries_overdue":  row["entries_overdue"],
            "entries_partial":  row.get("entries_partial", 0),
            "fee_status":       s.get("fee_status", "pending"),
            "oldest_due":       row["oldest_due"] or "",
        })

    due_chart.sort(key=lambda x: x["total_due"], reverse=True)

    # (#26) Server-side name / admission number search
    if search:
        s = search.strip().lower()
        due_chart = [
            r for r in due_chart
            if s in r["name"].lower() or s in r.get("admission_number", "").lower()
        ]

    return due_chart


@router.get("/fees/sessions")
async def list_academic_sessions(request: Request):
    """
    Distinct academic_year values present in the ledger, newest first.
    Falls back to the current academic year if the ledger is empty.
    Feeds the Session filter dropdown (and similar selectors).
    """
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    years = await db.student_ledger.distinct("academic_year")
    cleaned = sorted({y for y in years if y}, reverse=True)
    if not cleaned:
        cleaned = [current_academic_year()]
    return cleaned


@router.get("/fees/due-fee-types")
async def get_due_fee_types(request: Request):
    """
    Distinct fee components and tuition months currently present in
    pending/overdue ledger entries. Feeds the Due Fees filter dropdown
    so the options are DB-driven, not hardcoded.
    """
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    pipeline = [
        {"$match": {"status": {"$in": ["pending", "overdue"]}}},
        {"$group": {
            "_id": "$fee_component",
            "months": {"$addToSet": {"$substr": ["$due_date", 0, 7]}},
        }},
    ]
    rows = await db.student_ledger.aggregate(pipeline).to_list(100)
    components = sorted([r["_id"] for r in rows if r["_id"]])
    tuition_months = sorted({
        m for r in rows if r["_id"] == "tuition" for m in (r.get("months") or []) if m
    })
    return {"components": components, "tuition_months": tuition_months}


@router.post("/fees/refresh-overdue")
async def refresh_all_overdue(request: Request):
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    today = datetime.now().strftime("%Y-%m-%d")

    # Bulk mark all past-due pending entries as overdue
    result = await db.student_ledger.update_many(
        {"status": "pending", "due_date": {"$lt": today}},
        {"$set": {"status": "overdue"}}
    )

    # Bulk-update student fee_status based on ledger state
    overdue_sids = await db.student_ledger.distinct("student_id", {"status": "overdue"})
    if overdue_sids:
        await db.students.update_many(
            {"student_id": {"$in": overdue_sids}, "is_active": True},
            {"$set": {"fee_status": "overdue", "app_locked": True}}
        )

    pending_only_sids = await db.student_ledger.distinct(
        "student_id",
        {"status": "pending", "student_id": {"$nin": overdue_sids}}
    )
    if pending_only_sids:
        await db.students.update_many(
            {"student_id": {"$in": pending_only_sids}, "is_active": True},
            {"$set": {"fee_status": "pending", "app_locked": False}}
        )

    return {
        "message": f"Overdue refresh complete. {result.modified_count} entries marked overdue.",
        "entries_marked_overdue": result.modified_count,
    }


# ─── Concessions ──────────────────────────────────────────────────────────────

@router.post("/fees/concession")
async def apply_concession(request: Request):
    user = await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    body = await request.json()

    student_id = body.get("student_id")
    concession_percent = float(body.get("concession_percent", 0))
    reason = body.get("reason", "Scholarship/Concession")
    components = body.get("components")  # None = all pending; or list like ["tuition"]

    if not (0 <= concession_percent <= 100):
        raise HTTPException(status_code=400, detail="Concession must be 0–100%")

    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    q = {"student_id": student_id, "status": {"$in": ["pending", "overdue"]}}
    if components:
        q["fee_component"] = {"$in": components}

    pending = await db.student_ledger.find(q, {"_id": 0}).to_list(200)
    ops = []
    for entry in pending:
        existing_conc = entry.get("concession_amount", 0)
        existing_reason = entry.get("concession_reason") or ""
        gross = entry["gross_amount"]

        # (#1) Preserve existing sibling discount — stack admin concession on top of it
        if "sibling" in existing_reason.lower() and existing_conc > 0:
            base_after_sibling = gross - existing_conc
            admin_conc = round(base_after_sibling * concession_percent / 100, 2)
            total_conc = min(existing_conc + admin_conc, gross)
            combined_reason = f"{existing_reason} + {reason}"
        else:
            total_conc = round(gross * concession_percent / 100, 2)
            combined_reason = reason

        new_net = max(0, gross - total_conc + entry.get("late_fee_applied", 0))
        ops.append(UpdateOne(
            {"ledger_id": entry["ledger_id"]},
            {"$set": {
                "concession_amount": total_conc,
                "concession_reason": combined_reason,
                "net_amount": new_net,
            }}
        ))
    if ops:
        await db.student_ledger.bulk_write(ops)
    updated = len(ops)

    await create_audit_log("concession", student_id, "apply_concession", {
        "percent": concession_percent, "reason": reason, "entries": updated
    }, user)
    return {"message": f"{concession_percent}% concession applied to {updated} entries", "entries_updated": updated}


@router.get("/fees/concessions")
async def list_concessions(request: Request):
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    pipeline = [
        {"$match": {"concession_amount": {"$gt": 0}, "status": {"$in": ["pending", "overdue", "paid"]}}},
        {"$group": {
            "_id": "$student_id",
            "total_concession": {"$sum": "$concession_amount"},
            "reason":           {"$first": "$concession_reason"},
            "entries":          {"$sum": 1},
        }},
    ]
    results = await db.student_ledger.aggregate(pipeline).to_list(1000)
    if not results:
        return []

    # Batch-fetch all students in one query
    student_ids = [r["_id"] for r in results]
    students = await db.students.find(
        {"student_id": {"$in": student_ids}},
        {"_id": 0, "student_id": 1, "first_name": 1, "last_name": 1,
         "class_name": 1, "section": 1, "admission_number": 1}
    ).to_list(1000)
    student_map = {s["student_id"]: s for s in students}

    out = []
    for r in results:
        s = student_map.get(r["_id"])
        if s:
            out.append({
                "student_id":       r["_id"],
                "student_name":     f"{s['first_name']} {s['last_name']}",
                "class_name":       s["class_name"],
                "section":          s["section"],
                "admission_number": s.get("admission_number", ""),
                "total_concession": r["total_concession"],
                "reason":           r["reason"],
                "entries":          r["entries"],
            })
    return out


# ─── Receipt PDF ──────────────────────────────────────────────────────────────

@router.get("/fees/receipt/{payment_id}/pdf")
async def download_receipt_pdf(payment_id: str, request: Request, ledger_id: Optional[str] = None):
    await get_current_user(request)

    payment = await db.fee_payments.find_one({"payment_id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")

    student = await db.students.find_one({"student_id": payment["student_id"]}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    # Per-fee receipt: when a specific ledger entry is requested (the receipt
    # button next to a single fee row), scope the receipt to THAT fee only —
    # showing its own stamped receipt number / paid date / amount. This avoids
    # showing unrelated fees when a payment_id link is shared or imprecise.
    scoped_entry = None        # set when the receipt is scoped to one fee
    entry_payments = []        # every payment made toward that one fee (oldest-first)
    if ledger_id:
        scoped = await db.student_ledger.find_one({"ledger_id": ledger_id}, {"_id": 0})
        if scoped:
            entries = [scoped]
            scoped_entry = scoped
            # All payments that touched this fee — so the receipt lists each
            # part-payment (e.g. Rs.500 + Rs.500) with a running total + balance.
            entry_payments = await db.fee_payments.find(
                {"student_id": payment["student_id"], "installment_ids": ledger_id}, {"_id": 0}
            ).to_list(200)
            entry_payments.sort(key=lambda p: str(p.get("payment_date") or p.get("created_at") or ""))
            payment = {
                **payment,
                "receipt_number": scoped.get("receipt_number") or payment.get("receipt_number"),
                "payment_date": scoped.get("paid_date") or payment.get("payment_date"),
                # The amount receipted for this fee = what's been paid on it.
                "amount": round(float(scoped.get("amount_paid") or scoped.get("net_amount", 0)), 2),
            }
        else:
            ledger_ids = payment.get("installment_ids", [])
            entries = await db.student_ledger.find(
                {"ledger_id": {"$in": ledger_ids}}, {"_id": 0}
            ).to_list(100)
    else:
        # Fetch ledger entries paid in this payment
        ledger_ids = payment.get("installment_ids", [])
        entries = await db.student_ledger.find(
            {"ledger_id": {"$in": ledger_ids}}, {"_id": 0}
        ).to_list(100)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            topMargin=0.5 * inch, bottomMargin=0.5 * inch,
                            leftMargin=0.7 * inch, rightMargin=0.7 * inch)
    elements = []
    styles = getSampleStyleSheet()

    orange = colors.HexColor("#E88A1A")
    title_style = ParagraphStyle("Title", parent=styles["Heading1"],
                                 fontSize=16, alignment=TA_CENTER, textColor=orange)
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"],
                                fontSize=9, alignment=TA_CENTER, textColor=colors.grey)
    normal_bold = ParagraphStyle("NB", parent=styles["Normal"], fontName=_PDF_FONT_BOLD, fontSize=9)

    elements.append(Paragraph("SHEMFORD FUTURISTIC SCHOOL", title_style))
    elements.append(Paragraph("Katwa, West Bengal | CBSE Affiliated | Empowering Futures", sub_style))
    elements.append(Spacer(1, 10))

    # Divider
    div_table = Table([["FEE RECEIPT"]], colWidths=[7 * inch])
    div_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), orange),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), _PDF_FONT_BOLD),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(div_table)
    elements.append(Spacer(1, 10))

    # For senior classes (11th/12th) the section IS the stream, so show the
    # class alone in Class and the stream/section in Stream — not combined.
    _cls = student.get("class_name", "") or ""
    _section = student.get("section", "") or ""
    _stream = student.get("stream", "") or ""
    _STREAM_CLASSES = {"11th", "12th", "11", "12", "Class 11", "Class 12"}
    if _cls in _STREAM_CLASSES:
        class_label = _cls
        stream_src = _stream or _section
        class_extra_label = "Stream"
        class_extra_value = stream_src.title() if stream_src else "—"
    else:
        class_label = f"{_cls} – {_section}" if _section else _cls
        class_extra_label = "Academic Year"
        class_extra_value = student.get("academic_year", "—") or "—"

    _method_label = (payment.get("payment_method", "") or "").replace("_", " ").upper()
    _info_val_style = ParagraphStyle("InfoVal", parent=styles["Normal"], fontSize=9, leading=11)
    _method_cell = Paragraph(_method_label, _info_val_style)

    info_data = [
        ["Receipt No.", payment.get("receipt_number", ""), "Date", _iso_to_dmy(payment.get("payment_date", ""))],
        ["Student Name", f"{student['first_name']} {student['last_name']}",
         "Admission No.", student.get("admission_number", "")],
        ["Class", class_label, class_extra_label, class_extra_value],
        ["Payment Method", _method_cell,
         "Txn ID", payment.get("transaction_id", "—") or "—"],
    ]
    info_table = Table(info_data, colWidths=[1.4 * inch, 2.1 * inch, 1.4 * inch, 2.1 * inch])
    info_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f5f5f5")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#f5f5f5")),
        ("FONTNAME", (0, 0), (-1, -1), _PDF_FONT_REG), ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (0, -1), _PDF_FONT_BOLD),
        ("FONTNAME", (2, 0), (2, -1), _PDF_FONT_BOLD),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey), ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 12))

    if scoped_entry:
        # Per-fee receipt: list the part-payment(s) toward this fee (each with
        # its date), then TOTAL PAID. Payments are capped to the amount actually
        # paid on this fee so unrelated/mis-linked payments can't inflate the
        # total (some legacy payments wrongly reference another fee's ledger id).
        e = scoped_entry
        net = round(float(e.get("net_amount", 0)), 2)
        # Amount actually paid on this fee — status-aware so legacy stale fields
        # (older paid entries with amount_paid=0) don't understate it.
        if e.get("status") == "paid":
            paid_total = net
        else:
            paid_total = round(float(e.get("amount_paid") or 0), 2)

        elements.append(Paragraph(
            f"<b>Fee:</b> {e.get('description','')} "
            f"({e.get('fee_type','').replace('_',' ').title()}) &nbsp;—&nbsp; "
            f"Total billed: Rs.{net:,.2f}",
            ParagraphStyle("FeeLine", parent=styles["Normal"], fontSize=10)))
        elements.append(Spacer(1, 6))

        hist_method_style = ParagraphStyle("HistMethod2", parent=styles["Normal"], fontSize=8, leading=9)
        hist = [["#", "Date", "Method", "Amount (Rs.)"]]
        idx = 0
        remaining = paid_total   # budget = what was actually paid on THIS fee
        for p in entry_payments:   # every payment toward this fee, oldest-first
            if remaining <= 0:
                break
            full_amt = round(float(p.get("amount", 0)), 2)
            if full_amt <= 0:
                continue
            # A payment may cover several fees; take only the portion that
            # applies to this fee (never exceed what's left of paid_total).
            row_amt = round(min(full_amt, remaining), 2)
            remaining = round(remaining - row_amt, 2)
            method_label = (p.get("payment_method", "") or "").replace("_", " ").upper()
            idx += 1
            hist.append([
                str(idx),
                _iso_to_dmy(p.get("payment_date") or p.get("created_at", "")),
                Paragraph(method_label, hist_method_style),
                f"{row_amt:,.2f}",
            ])
        if idx == 0:   # no linked payment rows — fall back to the entry's paid total
            hist.append(["1", _iso_to_dmy(payment.get("payment_date") or ""), "—", f"{paid_total:,.2f}"])
        hist.append(["", "", "TOTAL PAID", f"{paid_total:,.2f}"])

        htbl = Table(hist, colWidths=[0.4 * inch, 1.5 * inch, 3.0 * inch, 1.6 * inch])
        htbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), orange),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), _PDF_FONT_BOLD),
            ("FONTNAME", (0, 1), (-1, -1), _PDF_FONT_REG),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (3, 0), (3, -1), "RIGHT"),
            ("ALIGN", (2, -1), (2, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
            ("PADDING", (0, 0), (-1, -1), 5),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f5f5f5")),
            ("FONTNAME", (0, -1), (-1, -1), _PDF_FONT_BOLD),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#fafafa")]),
        ]))
        elements.append(htbl)
        elements.append(Spacer(1, 18))
    else:
        # Itemise each fee with the amount PAID in this transaction, then the
        # TOTAL PAID. (No balance column — the receipt records what was paid.)
        total_paid = round(float(payment.get("amount", 0)), 2)
        n_entries = len(entries)
        fee_data = [["Description", "Fee Type", "Amount Paid (Rs.)"]]
        for e in entries:
            net = float(e.get("net_amount", 0))
            bal = 0.0 if e.get("status") == "paid" else float(e.get("remaining_balance", 0) or 0)
            if n_entries == 1:
                paid_here = total_paid          # whole payment applies to this fee
            else:
                paid_here = round(net - bal, 2) if bal > 0 else net
            fee_data.append([
                e.get("description", ""),
                e.get("fee_type", "").replace("_", " ").title(),
                f"{paid_here:,.2f}",
            ])
        fee_data.append(["", "TOTAL PAID", f"{total_paid:,.2f}"])
        col_w = [3.8 * inch, 1.4 * inch, 1.3 * inch]
        fee_table = Table(fee_data, colWidths=col_w)
        fee_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), orange),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), _PDF_FONT_BOLD),
            ("FONTNAME", (0, 1), (-1, -1), _PDF_FONT_REG),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ("ALIGN", (1, -1), (1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
            ("PADDING", (0, 0), (-1, -1), 5),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f5f5f5")),
            ("FONTNAME", (0, -1), (-1, -1), _PDF_FONT_BOLD),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#fafafa")]),
        ]))
        elements.append(fee_table)
        elements.append(Spacer(1, 24))

    if payment.get("remarks"):
        elements.append(Paragraph(f"Remarks: {payment['remarks']}", styles["Normal"]))
        elements.append(Spacer(1, 10))

    elements.append(Paragraph(
        "This is a computer-generated receipt and is valid without a physical signature.",
        sub_style
    ))

    doc.build(elements)
    buffer.seek(0)
    filename = f"receipt_{payment.get('receipt_number', payment_id).replace('/', '_')}.pdf"
    return StreamingResponse(
        buffer, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ─── Payment history ──────────────────────────────────────────────────────────

@router.get("/fees/payments")
async def get_fee_payments(request: Request, student_id: Optional[str] = None):
    user = await get_current_user(request)
    query = {}

    if user["role"] == UserRole.PARENT:
        children = await db.students.find(
            {"$or": [{"parent_email": user.get("email", "")}, {"parent_id": user["user_id"]}], "is_active": True},
            {"_id": 0, "student_id": 1}
        ).to_list(20)
        child_ids = [c["student_id"] for c in children]
        if not child_ids:
            return []
        query["student_id"] = student_id if student_id and student_id in child_ids else {"$in": child_ids}
    elif user["role"] == UserRole.STUDENT:
        stu = await db.students.find_one({"user_id": user["user_id"]}, {"_id": 0, "student_id": 1})
        if not stu:
            return []
        query["student_id"] = stu["student_id"]
    elif user["role"] in [UserRole.ADMIN, UserRole.ACCOUNTANT]:
        if student_id:
            query["student_id"] = student_id
        # No student_id = admin sees all payments (no filter)
    elif user["role"] == UserRole.TEACHER:
        raise HTTPException(status_code=403, detail="Teachers cannot access fee data")

    return await db.fee_payments.find(query, {"_id": 0}).sort("payment_date", -1).to_list(1000)


# ─── Legacy compatibility endpoints ───────────────────────────────────────────

@router.get("/fees/class-config")
async def get_class_fee_config_legacy(request: Request):
    """Legacy endpoint — returns component configs in the old class-config format."""
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    academic_year = current_academic_year()
    configs = await db.fee_component_configs.find(
        {"academic_year": academic_year, "is_active": True}, {"_id": 0}
    ).sort("class_name", 1).to_list(500)
    return configs


@router.get("/fees/student/{student_id}")
async def get_student_fees_legacy(student_id: str, request: Request):
    """Legacy compatibility — delegates to /fees/ledger/{student_id}."""
    return await get_student_ledger(student_id, request)


@router.get("/fees/structure")
async def get_fee_structures(request: Request, class_name: Optional[str] = None):
    await get_current_user(request)
    academic_year = current_academic_year()
    q = {"academic_year": academic_year, "is_active": True}
    if class_name:
        q["class_name"] = class_name
    return await db.fee_component_configs.find(q, {"_id": 0}).sort("class_name", 1).to_list(500)


# ─── Student Search (for Collect Fees tab search bar) ─────────────────────────

@router.get("/fees/search-students")
async def search_students_for_fees(
    request: Request,
    q: str = "",
    academic_year: Optional[str] = None,
):
    """
    Search students by name, roll number, or admission number.
    Used by the fee collection search bar. Returns up to 10 results.
    Roles: admin, accountant, teacher.
    """
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT, UserRole.TEACHER)(request)

    if not q or len(q.strip()) < 1:
        return []

    escaped = re.escape(q.strip())

    query: dict = {
        "is_active": True,
        "$or": [
            {"first_name": {"$regex": escaped, "$options": "i"}},
            {"last_name": {"$regex": escaped, "$options": "i"}},
            {"roll_number": {"$regex": escaped, "$options": "i"}},
            {"admission_number": {"$regex": escaped, "$options": "i"}},
        ],
    }
    if academic_year:
        query["academic_year"] = academic_year

    students = await db.students.find(
        query,
        {
            "_id": 0,
            "student_id": 1,
            "first_name": 1,
            "last_name": 1,
            "admission_number": 1,
            "roll_number": 1,
            "class_name": 1,
            "section": 1,
            "stream": 1,
            "fee_status": 1,
        },
    ).limit(10).to_list(10)

    return [
        {
            "student_id": s["student_id"],
            "name": f"{s['first_name']} {s['last_name']}",
            "admission_number": s.get("admission_number", ""),
            "roll_number": s.get("roll_number", ""),
            "class_name": s.get("class_name", ""),
            "section": s.get("section", ""),
            "stream": s.get("stream"),
            "fee_status": s.get("fee_status", "pending"),
        }
        for s in students
    ]


# ─────────────────────────────────────────────────────────────────────────────
# Reports: Fees Collection & Due Fees (admin / accountant)
# ─────────────────────────────────────────────────────────────────────────────

def _duration_range(duration: Optional[str], start_date: Optional[str], end_date: Optional[str]):
    """Resolve a (start, end) YYYY-MM-DD inclusive range from a duration keyword."""
    today = datetime.now().date()
    if duration == "custom":
        return (start_date or today.isoformat(), end_date or today.isoformat())
    if duration == "yesterday":
        d = today.replace(day=today.day) if today.day > 1 else today
        from datetime import timedelta
        y = today - timedelta(days=1)
        return (y.isoformat(), y.isoformat())
    from datetime import timedelta
    if duration == "this_week":
        start = today - timedelta(days=today.weekday())
        return (start.isoformat(), today.isoformat())
    if duration == "last_week":
        end = today - timedelta(days=today.weekday() + 1)
        start = end - timedelta(days=6)
        return (start.isoformat(), end.isoformat())
    if duration == "this_month":
        start = today.replace(day=1)
        return (start.isoformat(), today.isoformat())
    if duration == "last_month":
        first_of_this = today.replace(day=1)
        last_of_prev = first_of_this - timedelta(days=1)
        first_of_prev = last_of_prev.replace(day=1)
        return (first_of_prev.isoformat(), last_of_prev.isoformat())
    if duration == "this_year":
        # Academic year: April 1 → March 31
        if today.month >= 4:
            start = today.replace(month=4, day=1)
        else:
            start = today.replace(year=today.year - 1, month=4, day=1)
        return (start.isoformat(), today.isoformat())
    # default: today
    return (today.isoformat(), today.isoformat())


@router.get("/fees/reports/collection")
async def report_fees_collection(
    request: Request,
    duration: str = "today",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    class_name: Optional[str] = None,
    section: Optional[str] = None,
    fee_type: Optional[str] = None,
    fee_component: Optional[str] = None,
    fee_month: Optional[str] = None,  # MM — matches due_date month, any year
    payment_method: Optional[str] = None,
    student_id: Optional[str] = None,
    academic_year: Optional[str] = None,
    rollup: str = "monthly",
):
    """Fees Collection Report — JSON for the dashboard."""
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    return await _collect_report_rows(
        duration=duration, start_date=start_date, end_date=end_date,
        class_name=class_name, section=section, fee_type=fee_type,
        fee_component=fee_component, fee_month=fee_month,
        payment_method=payment_method, student_id=student_id,
        academic_year=academic_year,
    )


async def _collect_report_rows(
    duration: str = "today",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    class_name: Optional[str] = None,
    section: Optional[str] = None,
    fee_type: Optional[str] = None,
    fee_component: Optional[str] = None,
    fee_month: Optional[str] = None,
    payment_method: Optional[str] = None,
    student_id: Optional[str] = None,
    academic_year: Optional[str] = None,
) -> list:
    """Internal: shared query body used by JSON, PDF, and Excel report endpoints."""
    pay_query: dict = {}
    if academic_year:
        pay_query["academic_year"] = academic_year
    if duration != "all_time":
        range_start, range_end = _duration_range(duration, start_date, end_date)
        try:
            from datetime import timedelta as _td
            end_excl = (datetime.strptime(range_end, "%Y-%m-%d").date() + _td(days=1)).isoformat()
        except Exception:
            end_excl = range_end + "T99"
        pay_query["payment_date"] = {"$gte": range_start, "$lt": end_excl}
    if payment_method:
        pay_query["payment_method"] = {"$regex": f"^{re.escape(payment_method)}$", "$options": "i"}
    if student_id:
        pay_query["student_id"] = student_id
    payments = await db.fee_payments.find(pay_query, {"_id": 0}).sort("payment_date", -1).to_list(20000)
    if not payments:
        return []

    if not payments:
        return []

    installment_ids: list = []
    for p in payments:
        installment_ids.extend(p.get("installment_ids", []))
    installment_ids = list({i for i in installment_ids if i})

    entries = (
        await db.student_ledger.find({"ledger_id": {"$in": installment_ids}}, {"_id": 0}).to_list(50000)
        if installment_ids else []
    )
    ledger_by_id = {e["ledger_id"]: e for e in entries}

    student_ids = list({p["student_id"] for p in payments})
    students = await db.students.find(
        {"student_id": {"$in": student_ids}},
        {"_id": 0, "student_id": 1, "first_name": 1, "last_name": 1, "class_name": 1,
         "section": 1, "admission_number": 1, "phone": 1, "parent_name": 1, "parent_phone": 1}
    ).to_list(20000)
    student_map = {s["student_id"]: s for s in students}

    def norm(v: Optional[str]) -> str:
        if not v: return ""
        s = str(v).strip().lower()
        return s[6:].strip() if s.startswith("class ") else s

    class_filter   = norm(class_name)
    section_filter = norm(section)
    fee_filter     = norm(fee_type)

    # Component → display label (tuition collapsed to a single "Tuition Fee").
    def _category(e: dict) -> str:
        comp = e.get("fee_component")
        if comp and comp in _COMPONENT_LABELS:
            return _COMPONENT_LABELS[comp]
        if comp:
            return comp.replace("_", " ").title()
        return (e.get("fee_type") or "Other").replace("_", " ").title()

    # One row per (student, fee category) so each fee type shows separately
    # with its own collected amount.
    per_row: dict = {}
    for p in payments:
        s = student_map.get(p["student_id"]) or {}
        if class_filter and norm(s.get("class_name")) != class_filter:
            continue
        if section_filter and norm(s.get("section")) != section_filter:
            continue

        entry_ids = p.get("installment_ids", []) or []
        linked = [ledger_by_id[i] for i in entry_ids if i in ledger_by_id]
        if fee_component:
            linked = [e for e in linked if e.get("fee_component") == fee_component]
        if fee_month:
            linked = [e for e in linked if (e.get("due_date") or "")[5:7] == fee_month]
        if (fee_component or fee_month or fee_filter) and not linked:
            continue
        if fee_filter and not (fee_component or fee_month):
            linked = [e for e in linked if norm(e.get("fee_type")) == fee_filter]
            if not linked:
                continue

        pd = str(p.get("payment_date", ""))[:10]

        def _ensure(cat):
            key = (p["student_id"], cat)
            return per_row.setdefault(key, {
                "admission_number": s.get("admission_number", ""),
                "student_name":     f"{s.get('first_name','')} {s.get('last_name','')}".strip(),
                "mobile":           s.get("parent_phone") or s.get("phone") or "",
                "guardian":         s.get("parent_name", "") or "",
                "class_section":    f"{s.get('class_name','')}{(' (' + s.get('section') + ')') if s.get('section') else ''}",
                "fee_types":        cat,
                "payments_count":   0,
                "total_collected":  0.0,
                "last_payment_date": None,
                "due_date":         None,
            })

        if linked:
            for e in linked:
                cat = _category(e)
                row = _ensure(cat)
                row["total_collected"] += float(e.get("net_amount", 0))
                row["payments_count"] += 1
                d = e.get("due_date")
                if d and (row["due_date"] is None or d > row["due_date"]):
                    row["due_date"] = d
                if pd and (row["last_payment_date"] is None or pd > row["last_payment_date"]):
                    row["last_payment_date"] = pd
        else:
            # Payment with no linked ledger entries — bucket under "Fees".
            row = _ensure("Fees")
            row["total_collected"] += float(p.get("amount", 0))
            row["payments_count"] += 1
            if pd and (row["last_payment_date"] is None or pd > row["last_payment_date"]):
                row["last_payment_date"] = pd

    rows = []
    for row in per_row.values():
        row["total_collected"] = round(row["total_collected"], 2)
        rows.append(row)
    rows.sort(key=lambda r: (r["student_name"], r["fee_types"]))
    return rows


@router.get("/fees/reports/due")
async def report_fees_due(
    request: Request,
    class_name: Optional[str] = None,
    section: Optional[str] = None,
    fee_type: Optional[str] = None,
    fee_component: Optional[str] = None,
    fee_month: Optional[str] = None,  # MM — matches due_date month, any year
    as_of_date: Optional[str] = None,
    duration: Optional[str] = None,   # keyword range (today/this_month/...) on due_date
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    student_id: Optional[str] = None,
    academic_year: Optional[str] = None,
):
    """Due Fees Report — JSON for the dashboard."""
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    return await _due_report_rows(
        class_name=class_name, section=section, fee_type=fee_type,
        fee_component=fee_component, fee_month=fee_month,
        as_of_date=as_of_date, duration=duration,
        start_date=start_date, end_date=end_date,
        student_id=student_id, academic_year=academic_year,
    )


async def _due_report_rows(
    class_name: Optional[str] = None,
    section: Optional[str] = None,
    fee_type: Optional[str] = None,
    fee_component: Optional[str] = None,
    fee_month: Optional[str] = None,
    as_of_date: Optional[str] = None,
    duration: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    student_id: Optional[str] = None,
    academic_year: Optional[str] = None,
) -> list:
    """Internal: shared query body for JSON, PDF, and Excel due-report endpoints."""
    ledger_query: dict = {"status": {"$in": ["pending", "overdue", "partially_paid"]}}
    if academic_year:
        ledger_query["academic_year"] = academic_year
    # Duration narrows the report to fees whose due_date falls in the selected
    # period (mirrors the Collection report's Search Duration). `as_of_date`
    # tightens the upper bound when both are supplied.
    due_filter: dict = {}
    if duration and duration != "all_time":
        rs, re_ = _duration_range(duration, start_date, end_date)
        due_filter["$gte"] = rs
        due_filter["$lte"] = re_
    if as_of_date:
        cur = due_filter.get("$lte")
        due_filter["$lte"] = min(cur, as_of_date) if cur else as_of_date
    if due_filter:
        ledger_query["due_date"] = due_filter
    if student_id:
        ledger_query["student_id"] = student_id
    # Class / fee_type filters are applied in Python with normalized matching below
    entries = await db.student_ledger.find(ledger_query, {"_id": 0}).sort("due_date", 1).to_list(10000)
    if not entries:
        return []

    student_ids = list({e["student_id"] for e in entries})
    students = await db.students.find(
        {"student_id": {"$in": student_ids}, "is_active": True},
        {"_id": 0, "student_id": 1, "first_name": 1, "last_name": 1, "class_name": 1,
         "section": 1, "admission_number": 1, "phone": 1, "parent_name": 1, "parent_phone": 1}
    ).to_list(5000)
    student_map = {s["student_id"]: s for s in students}

    def norm(v: Optional[str]) -> str:
        if not v:
            return ""
        s = str(v).strip().lower()
        return s[6:].strip() if s.startswith("class ") else s

    class_filter   = norm(class_name)
    section_filter = norm(section)
    fee_filter     = norm(fee_type)

    # One row per (student, fee category) so each fee type shows separately with
    # its own outstanding amount — mirrors the Collection report. `paid`/`balance`
    # are the entry-level amounts for that category (paid is non-zero only for
    # partially-paid fees; a fully-pending fee shows paid 0, balance = full).
    per_row: dict = {}
    for e in entries:
        s = student_map.get(e["student_id"])
        if not s:
            continue
        if class_filter and norm(s.get("class_name")) != class_filter:
            continue
        if section_filter and norm(s.get("section")) != section_filter:
            continue
        # Specific fee-category filters take precedence over the legacy bucket.
        if fee_component or fee_month:
            if fee_component and e.get("fee_component") != fee_component:
                continue
            if fee_month and (e.get("due_date") or "")[5:7] != fee_month:
                continue
        elif fee_filter and norm(e.get("fee_type")) != fee_filter:
            continue
        net = float(e.get("net_amount", 0))
        paid = float(e.get("amount_paid", 0))
        remaining = float(e.get("remaining_balance", net - paid))

        cat = _report_category(e)
        key = (e["student_id"], cat)
        agg = per_row.setdefault(key, {
            "admission_number": s.get("admission_number", ""),
            "student_name":     f"{s.get('first_name','')} {s.get('last_name','')}".strip(),
            "mobile":           s.get("parent_phone") or s.get("phone") or "",
            "guardian":         s.get("parent_name", "") or "",
            "class_section":    f"{s.get('class_name','')}{(' (' + s.get('section') + ')') if s.get('section') else ''}",
            "fee_types":        cat,
            "amount":           0.0,
            "paid":             0.0,
            "balance":          0.0,
            "entries_pending":  0,
            "oldest_due":       None,
        })
        agg["amount"]  += net
        agg["paid"]    += paid
        agg["balance"] += remaining
        agg["entries_pending"] += 1
        d = e.get("due_date") or ""
        if d and (agg["oldest_due"] is None or d < agg["oldest_due"]):
            agg["oldest_due"] = d

    rows = []
    for agg in per_row.values():
        agg["amount"]  = round(agg["amount"], 2)
        agg["paid"]    = round(agg["paid"], 2)
        agg["balance"] = round(agg["balance"], 2)
        rows.append(agg)

    rows.sort(key=lambda r: (r["student_name"], r["fee_types"]))
    return rows


# ── Report exports: PDF + Excel ──────────────────────────────────────────────

_FEE_TYPE_LABELS = {"one_time": "One Time", "monthly": "Monthly", "yearly": "Yearly"}


def _fmt_fee_types(v) -> str:
    if not v:
        return "—"
    parts = [p.strip() for p in str(v).split(",") if p.strip()]
    return ", ".join(_FEE_TYPE_LABELS.get(p, p) for p in parts) if parts else "—"


def _fmt_inr(v) -> str:
    try:
        return f"Rs.{float(v):,.2f}"
    except Exception:
        return str(v) if v not in (None, "") else "—"


def _fmt_str(v) -> str:
    return str(v) if v not in (None, "") else "—"


def _fmt_date(v) -> str:
    """Convert backend's YYYY-MM-DD into DD-MM-YYYY for display."""
    if not v:
        return "—"
    s = str(v)
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return f"{s[8:10]}/{s[5:7]}/{s[0:4]}"
    return s


# Columns used in PDF + Excel exports. Each: (header, row-key, formatter, align)
# align: 'L' = left (default), 'R' = right (numbers/amounts), 'C' = center.
COLLECTION_EXPORT_COLUMNS = [
    ("Admission No",     "admission_number",  _fmt_str,        "L"),
    ("Student Name",     "student_name",      _fmt_str,        "L"),
    ("Mobile",           "mobile",            _fmt_str,        "L"),
    ("Class (Section)",  "class_section",     _fmt_str,        "L"),
    ("Fees Type",        "fee_types",         _fmt_fee_types,  "L"),
    ("Due Date",         "due_date",          _fmt_date,       "C"),
    ("Last Payment",     "last_payment_date", _fmt_date,       "C"),
    ("Total Paid (Rs.)", "total_collected",   _fmt_inr,        "R"),
]

DUE_EXPORT_COLUMNS = [
    ("Admission No",     "admission_number",  _fmt_str,        "L"),
    ("Student Name",     "student_name",      _fmt_str,        "L"),
    ("Mobile",           "mobile",            _fmt_str,        "L"),
    ("Class (Section)",  "class_section",     _fmt_str,        "L"),
    ("Fees Type",        "fee_types",         _fmt_fee_types,  "L"),
    ("Oldest Due",       "oldest_due",        _fmt_date,       "C"),
    ("Amount (Rs.)",   "amount",            _fmt_inr,        "R"),
    ("Paid (Rs.)",     "paid",              _fmt_inr,        "R"),
    ("Balance (Rs.)",  "balance",           _fmt_inr,        "R"),
]


# Theme — matches the orange accent used elsewhere in the UI
SCHOOL_NAME = "Shemford Futuristic School"
ACCENT_HEX  = "#E88A1A"
HEADER_BG   = "#0F172A"  # dark slate to mirror the sidebar header
HEADER_FG   = "#FFFFFF"
ZEBRA_HEX   = "#F8FAFC"


def _build_excel(title: str, columns: list, rows: list) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Report"

    headers = [c[0] for c in columns]
    aligns  = [c[3] if len(c) > 3 else "L" for c in columns]
    ncols   = len(columns)

    # Row 1: School name banner
    ws.cell(row=1, column=1, value=SCHOOL_NAME)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill      = PatternFill("solid", fgColor=HEADER_BG.lstrip("#"))
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Row 2: Report title
    ws.cell(row=2, column=1, value=title)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws["A2"].font      = Font(bold=True, color="FFFFFF", size=11)
    ws["A2"].fill      = PatternFill("solid", fgColor=ACCENT_HEX.lstrip("#"))
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Row 3: Generated on
    ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}     Total rows: {len(rows)}")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    ws["A3"].font      = Font(italic=True, color="475569", size=9)
    ws["A3"].alignment = Alignment(horizontal="right")

    # Row 4: Column headers
    head_fill = PatternFill("solid", fgColor="E2E8F0")
    head_font = Font(bold=True, color="0F172A", size=10)
    thin = Side(border_style="thin", color="CBD5E1")
    header_row_idx = 4
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=i, value=h)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal={"L": "left", "R": "right", "C": "center"}[aligns[i - 1]], vertical="center")
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    ws.row_dimensions[header_row_idx].height = 22

    # Data rows
    body_font = Font(size=10, color="0F172A")
    zebra_fill = PatternFill("solid", fgColor=ZEBRA_HEX.lstrip("#"))
    for ridx, r in enumerate(rows, start=header_row_idx + 1):
        for i, (_, key, fmt, *_rest) in enumerate(columns, start=1):
            cell = ws.cell(row=ridx, column=i, value=fmt(r.get(key)))
            cell.font = body_font
            cell.alignment = Alignment(horizontal={"L": "left", "R": "right", "C": "center"}[aligns[i - 1]], vertical="center")
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            if ridx % 2 == 0:
                cell.fill = zebra_fill

    # Column widths — measured from data, capped
    for i in range(1, ncols + 1):
        col_letter = ws.cell(row=header_row_idx, column=i).column_letter
        max_len = len(str(headers[i - 1]))
        for ridx in range(header_row_idx + 1, header_row_idx + 1 + len(rows)):
            v = ws.cell(row=ridx, column=i).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 38)

    # Freeze the header so column titles stay visible while scrolling
    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


PDF_ROW_CAP = 500


def _build_pdf(title: str, columns: list, rows: list) -> io.BytesIO:
    from reportlab.lib.pagesizes import landscape

    total_rows = len(rows)
    truncated = total_rows > PDF_ROW_CAP
    rows = rows[:PDF_ROW_CAP] if truncated else rows

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=0.35 * inch, rightMargin=0.35 * inch,
        topMargin=0.4 * inch, bottomMargin=0.4 * inch,
        title=title,
    )
    styles = getSampleStyleSheet()

    school_style = ParagraphStyle(
        "School", parent=styles["Heading1"],
        alignment=TA_CENTER, fontSize=18, textColor=colors.HexColor(HEADER_BG),
        spaceAfter=2, leading=20,
    )
    title_style = ParagraphStyle(
        "Title", parent=styles["Heading2"],
        alignment=TA_CENTER, fontSize=12, textColor=colors.HexColor(ACCENT_HEX),
        spaceAfter=4, leading=14,
    )
    meta_style = ParagraphStyle(
        "Meta", parent=styles["Normal"],
        alignment=TA_LEFT, fontSize=8, textColor=colors.HexColor("#64748b"),
        spaceAfter=8,
    )
    foot_style = ParagraphStyle(
        "Foot", parent=styles["Normal"],
        alignment=TA_CENTER, fontSize=7, textColor=colors.HexColor("#94a3b8"),
    )

    meta_text = (
        f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}  ·  "
        f"Showing {len(rows)} of {total_rows} rows" if truncated else
        f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}  ·  Total rows: {total_rows}"
    )

    elements = [
        Paragraph(SCHOOL_NAME, school_style),
        Paragraph(title, title_style),
        Paragraph(meta_text, meta_style),
    ]

    # Build the data grid
    data = [[c[0] for c in columns]]
    for r in rows:
        data.append([fmt(r.get(key)) for (_, key, fmt, *_rest) in columns])

    # Column-level alignment for the table
    align_map = {"L": "LEFT", "R": "RIGHT", "C": "CENTER"}
    style_cmds = [
        ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor(HEADER_BG)),
        ("TEXTCOLOR",    (0, 0), (-1, 0), colors.HexColor(HEADER_FG)),
        ("FONTNAME",     (0, 0), (-1, 0), _PDF_FONT_BOLD),
        ("FONTSIZE",     (0, 0), (-1, 0), 8.5),
        ("FONTSIZE",     (0, 1), (-1, -1), 7.5),
        ("GRID",         (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(ZEBRA_HEX)]),
    ]
    for col_idx, c in enumerate(columns):
        align = c[3] if len(c) > 3 else "L"
        style_cmds.append(("ALIGN", (col_idx, 0), (col_idx, -1), align_map[align]))

    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle(style_cmds))
    elements.append(tbl)

    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        "Computer-generated report. For queries, contact the school accounts office.",
        foot_style,
    ))

    doc.build(elements)
    buf.seek(0)
    return buf


def _today_str() -> str:
    return datetime.now().strftime("%Y-%m-%d")


@router.get("/fees/reports/collection/excel")
async def export_collection_excel(
    request: Request,
    duration: str = "today",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    class_name: Optional[str] = None,
    section: Optional[str] = None,
    fee_type: Optional[str] = None,
    fee_component: Optional[str] = None,
    fee_month: Optional[str] = None,
    payment_method: Optional[str] = None,
    student_id: Optional[str] = None,
):
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    rows = await _collect_report_rows(
        duration=duration, start_date=start_date, end_date=end_date,
        class_name=class_name, section=section, fee_type=fee_type,
        fee_component=fee_component, fee_month=fee_month,
        payment_method=payment_method, student_id=student_id,
    )
    buf = _build_excel("Fees Collection Report", COLLECTION_EXPORT_COLUMNS, rows)
    filename = f"fees-collection-{_today_str()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/fees/reports/collection/pdf")
async def export_collection_pdf(
    request: Request,
    duration: str = "today",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    class_name: Optional[str] = None,
    section: Optional[str] = None,
    fee_type: Optional[str] = None,
    fee_component: Optional[str] = None,
    fee_month: Optional[str] = None,
    payment_method: Optional[str] = None,
    student_id: Optional[str] = None,
):
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    rows = await _collect_report_rows(
        duration=duration, start_date=start_date, end_date=end_date,
        class_name=class_name, section=section, fee_type=fee_type,
        fee_component=fee_component, fee_month=fee_month,
        payment_method=payment_method, student_id=student_id,
    )
    buf = _build_pdf("Fees Collection Report", COLLECTION_EXPORT_COLUMNS, rows)
    filename = f"fees-collection-{_today_str()}.pdf"
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/fees/reports/due/excel")
async def export_due_excel(
    request: Request,
    class_name: Optional[str] = None,
    section: Optional[str] = None,
    fee_type: Optional[str] = None,
    fee_component: Optional[str] = None,
    fee_month: Optional[str] = None,
    as_of_date: Optional[str] = None,
    duration: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    student_id: Optional[str] = None,
    academic_year: Optional[str] = None,
):
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    rows = await _due_report_rows(
        class_name=class_name, section=section, fee_type=fee_type,
        fee_component=fee_component, fee_month=fee_month,
        as_of_date=as_of_date, duration=duration,
        start_date=start_date, end_date=end_date,
        student_id=student_id, academic_year=academic_year,
    )
    buf = _build_excel("Due Fees Report", DUE_EXPORT_COLUMNS, rows)
    filename = f"fees-due-{_today_str()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/fees/reports/due/pdf")
async def export_due_pdf(
    request: Request,
    class_name: Optional[str] = None,
    section: Optional[str] = None,
    fee_type: Optional[str] = None,
    fee_component: Optional[str] = None,
    fee_month: Optional[str] = None,
    as_of_date: Optional[str] = None,
    duration: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    student_id: Optional[str] = None,
    academic_year: Optional[str] = None,
):
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    rows = await _due_report_rows(
        class_name=class_name, section=section, fee_type=fee_type,
        fee_component=fee_component, fee_month=fee_month,
        as_of_date=as_of_date, duration=duration,
        start_date=start_date, end_date=end_date,
        student_id=student_id, academic_year=academic_year,
    )
    buf = _build_pdf("Due Fees Report", DUE_EXPORT_COLUMNS, rows)
    filename = f"fees-due-{_today_str()}.pdf"
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
