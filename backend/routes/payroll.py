"""
Shemford School — Payroll System
=================================

Salary calculation rules:
  1. Full month:       net = monthly_salary - lwp_deduction - other_deductions
  2. Mid-month join:   gross = (working_days / total_days) * monthly_salary
  3. LWP deduction:    per_day = monthly_salary / total_days; deduction = lwp_days * per_day

Data integrity:
  - Unique index on (employee_id, month_year) prevents duplicate generation.
  - Salary snapshot frozen at generation time — historical records survive salary changes.
  - All financial mutations audit-logged.

Exports:
  - Excel: openpyxl, bank-ready format (name, phone, address, account, IFSC, amount)
  - PDF:   reportlab (payslip, Form 16, yearly statement)
"""

import calendar
import io
import logging
import uuid
from datetime import datetime, timezone, date
from typing import List, Optional

from fastapi import APIRouter, BackgroundTasks, HTTPException, Query, Request, Response
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
)

from database import db
from models import UserRole, PayrollRecord, PayrollStatus
from auth_utils import get_current_user, require_roles, create_audit_log, get_rid, ensure_active_session
from security import decrypt_bank_fields

router = APIRouter()
logger = logging.getLogger(__name__)

SCHOOL_NAME    = "Shemford Futuristic School"
SCHOOL_ADDRESS = "Shemford School Campus"

# ── Request schemas ───────────────────────────────────────────────────────────

class GeneratePayrollRequest(BaseModel):
    month: int                              # 1–12
    year: int
    employee_ids: Optional[List[str]] = None  # None = all active employees
    overwrite: bool = False                 # allow re-generation (admin override)


class UpdatePayrollRequest(BaseModel):
    other_deductions: Optional[float] = None
    deduction_remarks: Optional[str] = None
    lwp_days: Optional[float] = None
    remarks: Optional[str] = None


class MarkPaidRequest(BaseModel):
    payment_reference: Optional[str] = None


# ── Salary calculation engine ─────────────────────────────────────────────────

def calculate_payroll(
    employee: dict,
    month: int,
    year: int,
    lwp_days: float = 0.0,
    other_deductions: float = 0.0,
) -> dict:
    """
    Pure function — no I/O.
    Returns a dict with all calculated fields.
    """
    total_days    = calendar.monthrange(year, month)[1]
    monthly_salary = float(employee.get("monthly_salary") or employee.get("salary") or 0)
    per_day       = monthly_salary / total_days if total_days else 0

    # Mid-month joining check
    joining_date_str = employee.get("joining_date", "")
    is_mid_month     = False
    working_days     = total_days

    if joining_date_str:
        try:
            jd = datetime.strptime(joining_date_str, "%Y-%m-%d").date()
            if jd.year == year and jd.month == month and jd.day > 1:
                is_mid_month = True
                working_days = total_days - jd.day + 1
        except ValueError:
            pass

    # Gross = prorated if mid-month, else full monthly salary
    gross = (working_days / total_days) * monthly_salary if total_days else 0

    # LWP cannot exceed working days
    lwp_days      = min(float(lwp_days), working_days)
    lwp_deduction = round(lwp_days * per_day, 2)
    other_deductions = round(float(other_deductions), 2)
    total_deductions = round(lwp_deduction + other_deductions, 2)
    net_salary       = round(gross - total_deductions, 2)
    present_days     = working_days - lwp_days

    return {
        "total_days":       total_days,
        "working_days":     working_days,
        "lwp_days":         lwp_days,
        "present_days":     present_days,
        "per_day_salary":   round(per_day, 2),
        "monthly_salary":   round(monthly_salary, 2),
        "gross_salary":     round(gross, 2),
        "lwp_deduction":    lwp_deduction,
        "other_deductions": other_deductions,
        "total_deductions": total_deductions,
        "net_salary":       net_salary,
        "is_mid_month_join": is_mid_month,
    }


# ── POST /payroll/generate ────────────────────────────────────────────────────

@router.post("/payroll/generate")
async def generate_payroll(body: GeneratePayrollRequest, request: Request, background_tasks: BackgroundTasks):
    """
    Generate payroll for one month.
    - If employee_ids is None → generate for all active employees.
    - Skips employees who already have a payroll for the month (unless overwrite=True).
    - Returns summary of generated / skipped / failed records.
    """
    user = await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)
    await ensure_active_session(request)  # previous sessions are read-only

    if not (1 <= body.month <= 12):
        raise HTTPException(status_code=400, detail="Month must be between 1 and 12.")
    if body.year < 2000 or body.year > 2100:
        raise HTTPException(status_code=400, detail="Invalid year.")

    month_year = f"{body.year}-{body.month:02d}"

    # Fetch employees
    query: dict = {"is_active": True}
    if body.employee_ids:
        query["employee_id"] = {"$in": body.employee_ids}
    employees = await db.employees.find(query, {"_id": 0}).to_list(1000)

    if not employees:
        raise HTTPException(status_code=404, detail="No active employees found.")

    generated, skipped, failed = [], [], []

    for emp in employees:
        emp = decrypt_bank_fields(emp)
        eid = emp["employee_id"]
        existing = await db.payroll.find_one({"employee_id": eid, "month_year": month_year})

        if existing and not body.overwrite:
            skipped.append({"employee_id": eid, "reason": "already_generated"})
            continue

        try:
            # Warn if employee is missing bank details (cannot disburse salary)
            missing_bank = []
            if not emp.get("bank_account_number"):
                missing_bank.append("bank_account_number")
            if not emp.get("bank_ifsc"):
                missing_bank.append("bank_ifsc")
            if not emp.get("bank_account_holder"):
                missing_bank.append("bank_account_holder")
            if missing_bank:
                logger.warning(
                    "Employee %s is missing bank fields: %s — payroll generated but bank transfer not possible",
                    eid, ", ".join(missing_bank)
                )

            calc = calculate_payroll(emp, body.month, body.year)

            record = PayrollRecord(
                employee_id=eid,
                month=body.month,
                year=body.year,
                month_year=month_year,
                generated_by=user["user_id"],
                bank_account_number=emp.get("bank_account_number"),
                bank_ifsc=emp.get("bank_ifsc"),
                bank_name=emp.get("bank_name"),
                **calc,
            )
            doc = record.model_dump()
            doc["created_at"] = doc["created_at"].isoformat()
            doc["updated_at"] = doc["updated_at"].isoformat()

            if existing and body.overwrite:
                doc.pop("payroll_id", None)
                await db.payroll.replace_one(
                    {"employee_id": eid, "month_year": month_year}, doc, upsert=True
                )
                entry = {"employee_id": eid, "payroll_id": existing["payroll_id"], "action": "overwritten"}
            else:
                await db.payroll.insert_one(doc)
                entry = {"employee_id": eid, "payroll_id": record.payroll_id, "action": "created"}

            if missing_bank:
                entry["warnings"] = [f"Missing bank field(s): {', '.join(missing_bank)}. Bank transfer disabled."]
            generated.append(entry)

        except Exception as exc:
            logger.error("Payroll generation failed for %s: %s", eid, exc)
            failed.append({"employee_id": eid, "error": str(exc)})

    background_tasks.add_task(
        create_audit_log,
        "payroll", month_year, "generate",
        {
            "month_year": month_year,
            "generated": len(generated),
            "skipped": len(skipped),
            "failed": len(failed),
            "generated_by_role": user["role"],
        },
        user,
    )
    logger.info(
        "Payroll generated for %s: generated=%d skipped=%d failed=%d by=%s rid=%s",
        month_year, len(generated), len(skipped), len(failed), user["user_id"], get_rid(request)
    )

    return {
        "month_year": month_year,
        "generated": generated,
        "skipped": skipped,
        "failed": failed,
        "summary": {"generated": len(generated), "skipped": len(skipped), "failed": len(failed)},
    }


# ── GET /payroll ──────────────────────────────────────────────────────────────

@router.get("/payroll")
async def list_payroll(
    request: Request,
    response: Response,
    month_year: Optional[str] = None,   # "YYYY-MM"
    year: Optional[int] = None,
    status: Optional[str] = None,
    employee_id: Optional[str] = None,
    limit: int = Query(default=30, le=200),
    page: int = Query(default=1, ge=1),
):
    """Admin/Accountant: full payroll list. Teacher: own records only."""
    user = await get_current_user(request)

    if user["role"] not in (UserRole.ADMIN, UserRole.ACCOUNTANT, UserRole.TEACHER):
        raise HTTPException(status_code=403, detail="Access denied. Payroll is restricted to staff only.")

    query: dict = {}

    # ── TEACHER: strictly filter to their own employee record ────────────────────
    if user["role"] == UserRole.TEACHER:
        emp = await db.employees.find_one(
            {"$or": [{"user_id": user["user_id"]}, {"email": user["email"]}]},
            {"employee_id": 1}
        )
        if not emp:
            logger.warning("Payroll list requested by teacher %s with no linked employee record", user["user_id"])
            return []
        # Ignore any employee_id query param from teacher — always force their own ID
        query["employee_id"] = emp["employee_id"]
        logger.info("Payroll list accessed by employee %s (user_id=%s)", emp["employee_id"], user["user_id"])
    else:
        # Admin / Accountant: may optionally filter by employee
        if employee_id:
            query["employee_id"] = employee_id
        logger.info("Payroll list accessed by %s (role=%s)", user["user_id"], user["role"])

    if month_year:
        query["month_year"] = month_year
    elif year:
        query["year"] = year
    if status:
        query["status"] = status

    import asyncio as _asyncio
    skip = (page - 1) * limit
    total, records = await _asyncio.gather(
        db.payroll.count_documents(query),
        db.payroll.find(query, {"_id": 0}).sort("month_year", -1).skip(skip).limit(limit).to_list(limit),
    )
    pages = max(1, -(-total // limit))
    response.headers["X-Total-Count"] = str(total)
    response.headers["X-Total-Pages"] = str(pages)
    response.headers["X-Page"] = str(page)

    # Enrich with employee name (Admin/Accountant view only — teachers see their own name naturally)
    emp_ids = list({r["employee_id"] for r in records})
    emps = await db.employees.find(
        {"employee_id": {"$in": emp_ids}},
        {"_id": 0, "employee_id": 1, "first_name": 1, "last_name": 1, "designation": 1, "department": 1}
    ).to_list(len(emp_ids))
    emp_map = {e["employee_id"]: e for e in emps}

    for r in records:
        emp = emp_map.get(r["employee_id"], {})
        r["employee_name"] = f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip()
        r["designation"]   = emp.get("designation", "")
        r["department"]    = emp.get("department", "")

    return records


# ── GET /payroll/{payroll_id} ─────────────────────────────────────────────────

@router.get("/payroll/{payroll_id}")
async def get_payroll_record(payroll_id: str, request: Request):
    user = await get_current_user(request)

    # Only staff roles may access payroll records
    if user["role"] not in (UserRole.ADMIN, UserRole.ACCOUNTANT, UserRole.TEACHER):
        raise HTTPException(status_code=403, detail="Access denied. Payroll records are restricted to staff.")

    record = await db.payroll.find_one({"payroll_id": payroll_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Payroll record not found.")

    # TEACHER: enforce strict own-record access
    if user["role"] == UserRole.TEACHER:
        emp = await db.employees.find_one(
            {"$or": [{"user_id": user["user_id"]}, {"email": user["email"]}]},
            {"employee_id": 1}
        )
        if not emp or emp["employee_id"] != record["employee_id"]:
            logger.warning(
                "Unauthorized payroll access: user=%s tried to read payroll=%s (owner=%s)",
                user["user_id"], payroll_id, record["employee_id"]
            )
            raise HTTPException(status_code=403, detail="Access denied. You can only view your own payroll.")
        logger.info("Payroll %s accessed by employee %s", payroll_id, emp["employee_id"])
    else:
        logger.info("Payroll %s accessed by %s (role=%s)", payroll_id, user["user_id"], user["role"])

    emp = await db.employees.find_one({"employee_id": record["employee_id"]}, {"_id": 0})
    if emp:
        emp_data = decrypt_bank_fields(emp)
        # Strip sensitive bank fields from non-admin responses for the teacher viewing their own slip
        if user["role"] == UserRole.TEACHER:
            emp_data.pop("bank_account_number", None)
            emp_data.pop("bank_ifsc", None)
        record["employee"] = emp_data

    return record


# ── PATCH /payroll/{payroll_id} ───────────────────────────────────────────────

@router.patch("/payroll/{payroll_id}")
async def update_payroll(payroll_id: str, body: UpdatePayrollRequest, request: Request):
    """
    Admin can update LWP days or other deductions before approving.
    Recalculates net salary automatically.
    """
    user = await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)

    record = await db.payroll.find_one({"payroll_id": payroll_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Payroll record not found.")
    if record["status"] == PayrollStatus.PAID:
        raise HTTPException(status_code=400, detail="Cannot edit a paid payroll record.")
    if record["status"] == PayrollStatus.APPROVED:
        raise HTTPException(
            status_code=400,
            detail="Payroll is locked after approval. Revert to draft status before editing."
        )

    updates: dict = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if body.other_deductions is not None:
        updates["other_deductions"] = round(float(body.other_deductions), 2)
    if body.deduction_remarks is not None:
        updates["deduction_remarks"] = body.deduction_remarks
    if body.lwp_days is not None:
        updates["lwp_days"] = max(0.0, float(body.lwp_days))
    if body.remarks is not None:
        updates["remarks"] = body.remarks

    # Recalculate if financial fields changed
    if "other_deductions" in updates or "lwp_days" in updates:
        lwp    = updates.get("lwp_days",         record["lwp_days"])
        other  = updates.get("other_deductions", record["other_deductions"])
        per_day = record["per_day_salary"]

        lwp_deduction    = round(float(lwp) * per_day, 2)
        total_deductions = round(lwp_deduction + float(other), 2)
        net_salary       = round(record["gross_salary"] - total_deductions, 2)
        present_days     = record["working_days"] - float(lwp)

        updates.update({
            "lwp_deduction":    lwp_deduction,
            "total_deductions": total_deductions,
            "net_salary":       net_salary,
            "present_days":     present_days,
        })

    await db.payroll.update_one({"payroll_id": payroll_id}, {"$set": updates})
    await create_audit_log("payroll", payroll_id, "update", updates, user)
    return await db.payroll.find_one({"payroll_id": payroll_id}, {"_id": 0})


# ── POST /payroll/{payroll_id}/approve ────────────────────────────────────────

@router.post("/payroll/{payroll_id}/approve")
async def approve_payroll(payroll_id: str, request: Request):
    user = await require_roles(UserRole.ADMIN)(request)

    record = await db.payroll.find_one({"payroll_id": payroll_id})
    if not record:
        raise HTTPException(status_code=404, detail="Payroll record not found.")
    if record["status"] == PayrollStatus.APPROVED:
        return {"message": "Already approved.", "payroll_id": payroll_id}
    if record["status"] == PayrollStatus.PAID:
        raise HTTPException(status_code=400, detail="Payroll is already paid.")

    now = datetime.now(timezone.utc).isoformat()
    await db.payroll.update_one(
        {"payroll_id": payroll_id},
        {"$set": {"status": PayrollStatus.APPROVED, "approved_by": user["user_id"],
                  "approved_at": now, "updated_at": now}},
    )
    await create_audit_log("payroll", payroll_id, "approve", {}, user)
    return {"message": "Payroll approved.", "payroll_id": payroll_id}


# ── POST /payroll/{payroll_id}/revert-draft ──────────────────────────────────

@router.post("/payroll/{payroll_id}/revert-draft")
async def revert_payroll_to_draft(payroll_id: str, request: Request):
    """Admin only: revert an APPROVED payroll back to DRAFT so it can be edited."""
    user = await require_roles(UserRole.ADMIN)(request)

    record = await db.payroll.find_one({"payroll_id": payroll_id})
    if not record:
        raise HTTPException(status_code=404, detail="Payroll record not found.")
    if record["status"] == PayrollStatus.PAID:
        raise HTTPException(status_code=400, detail="Cannot revert a paid payroll record.")
    if record["status"] == PayrollStatus.DRAFT:
        return {"message": "Already in draft status.", "payroll_id": payroll_id}

    now = datetime.now(timezone.utc).isoformat()
    await db.payroll.update_one(
        {"payroll_id": payroll_id},
        {"$set": {"status": PayrollStatus.DRAFT, "approved_by": None,
                  "approved_at": None, "updated_at": now}},
    )
    await create_audit_log("payroll", payroll_id, "revert-draft", {}, user)
    logger.info("Payroll %s reverted to draft by %s", payroll_id, user["user_id"])
    return {"message": "Payroll reverted to draft.", "payroll_id": payroll_id}


# ── POST /payroll/{payroll_id}/mark-paid ─────────────────────────────────────

@router.post("/payroll/{payroll_id}/mark-paid")
async def mark_payroll_paid(payroll_id: str, body: MarkPaidRequest, request: Request):
    user = await require_roles(UserRole.ADMIN)(request)

    record = await db.payroll.find_one({"payroll_id": payroll_id})
    if not record:
        raise HTTPException(status_code=404, detail="Payroll record not found.")
    if record["status"] == PayrollStatus.PAID:
        return {"message": "Already marked as paid.", "payroll_id": payroll_id}
    if record["status"] != PayrollStatus.APPROVED:
        raise HTTPException(status_code=400, detail="Payroll must be approved before marking as paid.")

    now = datetime.now(timezone.utc).isoformat()
    await db.payroll.update_one(
        {"payroll_id": payroll_id},
        {"$set": {"status": PayrollStatus.PAID, "paid_at": now,
                  "payment_reference": body.payment_reference, "updated_at": now}},
    )
    await create_audit_log("payroll", payroll_id, "mark-paid",
                           {"payment_reference": body.payment_reference,
                            "employee_id": record["employee_id"],
                            "month_year": record["month_year"],
                            "net_salary": record["net_salary"]}, user)
    logger.info(
        "Payroll PAID: payroll=%s employee=%s month=%s net=Rs.%.2f ref=%s by=%s",
        payroll_id, record["employee_id"], record["month_year"],
        record["net_salary"], body.payment_reference, user["user_id"]
    )
    return {"message": "Payroll marked as paid.", "payroll_id": payroll_id}


# ── GET /payroll/employee/{employee_id} ───────────────────────────────────────

@router.get("/payroll/employee/{employee_id}")
async def get_employee_payroll(
    employee_id: str,
    request: Request,
    year: Optional[int] = None,
    limit: int = Query(default=24, le=120),
):
    """Employee can fetch their own payroll history; Admin/Accountant can fetch any."""
    user = await get_current_user(request)

    if user["role"] not in (UserRole.ADMIN, UserRole.ACCOUNTANT, UserRole.TEACHER):
        raise HTTPException(status_code=403, detail="Access denied. Payroll is restricted to staff only.")

    if user["role"] == UserRole.TEACHER:
        emp = await db.employees.find_one(
            {"$or": [{"user_id": user["user_id"]}, {"email": user["email"]}]},
            {"employee_id": 1}
        )
        if not emp:
            raise HTTPException(status_code=404, detail="No employee record linked to your account.")
        if emp["employee_id"] != employee_id:
            logger.warning(
                "Unauthorized payroll history access: user=%s tried to access employee=%s",
                user["user_id"], employee_id
            )
            raise HTTPException(status_code=403, detail="Access denied. You can only view your own payroll history.")
        logger.info("Payroll history accessed by employee %s", emp["employee_id"])

    # Verify employee exists
    target_emp = await db.employees.find_one({"employee_id": employee_id}, {"_id": 0, "first_name": 1, "last_name": 1})
    if not target_emp:
        raise HTTPException(status_code=404, detail="Employee not found.")

    query: dict = {"employee_id": employee_id}
    if year:
        query["year"] = year

    records = await db.payroll.find(query, {"_id": 0}).sort(
        [("year", -1), ("month", -1)]
    ).limit(limit).to_list(limit)

    return {
        "employee_id": employee_id,
        "employee_name": f"{target_emp['first_name']} {target_emp['last_name']}".strip(),
        "records": records,
        "total": len(records),
    }


# ── PDF: Payslip ──────────────────────────────────────────────────────────────

@router.get("/payroll/{payroll_id}/payslip")
async def download_payslip(payroll_id: str, request: Request):
    user = await get_current_user(request)

    if user["role"] not in (UserRole.ADMIN, UserRole.ACCOUNTANT, UserRole.TEACHER):
        raise HTTPException(status_code=403, detail="Access denied. Payslips are restricted to staff only.")

    record = await db.payroll.find_one({"payroll_id": payroll_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Payroll record not found.")

    # Access control — teacher can only download their own payslip
    if user["role"] == UserRole.TEACHER:
        emp = await db.employees.find_one(
            {"$or": [{"user_id": user["user_id"]}, {"email": user["email"]}]},
            {"employee_id": 1}
        )
        if not emp or emp["employee_id"] != record["employee_id"]:
            logger.warning(
                "Unauthorized payslip download: user=%s tried to download payroll=%s (owner=%s)",
                user["user_id"], payroll_id, record["employee_id"]
            )
            raise HTTPException(status_code=403, detail="Access denied. You can only download your own payslip.")
        logger.info("Payslip downloaded by employee %s for payroll %s", emp["employee_id"], payroll_id)

    emp = await db.employees.find_one({"employee_id": record["employee_id"]}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found.")
    emp = decrypt_bank_fields(emp)

    pdf_bytes = _generate_payslip_pdf(record, emp)
    month_name = calendar.month_name[record["month"]]
    filename = f"Payslip_{emp['first_name']}_{emp['last_name']}_{month_name}_{record['year']}.pdf"

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── PDF: Yearly Statement ─────────────────────────────────────────────────────

@router.get("/payroll/employee/{employee_id}/yearly-statement/{year}")
async def download_yearly_statement(employee_id: str, year: int, request: Request):
    user = await get_current_user(request)

    if user["role"] not in (UserRole.ADMIN, UserRole.ACCOUNTANT, UserRole.TEACHER):
        raise HTTPException(status_code=403, detail="Access denied. Restricted to staff only.")

    if user["role"] == UserRole.TEACHER:
        emp_check = await db.employees.find_one(
            {"$or": [{"user_id": user["user_id"]}, {"email": user["email"]}]},
            {"employee_id": 1}
        )
        if not emp_check or emp_check["employee_id"] != employee_id:
            raise HTTPException(status_code=403, detail="Access denied. You can only download your own statement.")

    records = await db.payroll.find(
        {"employee_id": employee_id, "year": year}, {"_id": 0}
    ).sort("month", 1).to_list(12)

    if not records:
        raise HTTPException(status_code=404, detail=f"No payroll records for year {year}.")

    emp = await db.employees.find_one({"employee_id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found.")
    emp = decrypt_bank_fields(emp)

    pdf_bytes = _generate_yearly_statement_pdf(records, emp, year)
    filename  = f"Yearly_Statement_{emp['first_name']}_{emp['last_name']}_{year}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── PDF: Form 16 ──────────────────────────────────────────────────────────────

@router.get("/payroll/employee/{employee_id}/form16/{year}")
async def download_form16(employee_id: str, year: int, request: Request):
    user = await get_current_user(request)

    if user["role"] not in (UserRole.ADMIN, UserRole.ACCOUNTANT, UserRole.TEACHER):
        raise HTTPException(status_code=403, detail="Access denied. Restricted to staff only.")

    if user["role"] == UserRole.TEACHER:
        emp_check = await db.employees.find_one(
            {"$or": [{"user_id": user["user_id"]}, {"email": user["email"]}]},
            {"employee_id": 1}
        )
        if not emp_check or emp_check["employee_id"] != employee_id:
            raise HTTPException(status_code=403, detail="Access denied. You can only download your own Form 16.")

    # Financial year: Apr (year-1) to Mar (year) → for FY 2025-26, year=2026
    fy_start_year = year - 1
    fy_end_year   = year

    records = await db.payroll.find(
        {
            "employee_id": employee_id,
            "$or": [
                {"year": fy_start_year, "month": {"$gte": 4}},   # Apr–Dec of start year
                {"year": fy_end_year,   "month": {"$lte": 3}},   # Jan–Mar of end year
            ],
            "status": PayrollStatus.PAID,
        },
        {"_id": 0}
    ).sort([("year", 1), ("month", 1)]).to_list(12)

    emp = await db.employees.find_one({"employee_id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found.")
    emp = decrypt_bank_fields(emp)

    pdf_bytes = _generate_form16_pdf(records, emp, fy_start_year, fy_end_year)
    filename  = f"Form16_{emp['first_name']}_{emp['last_name']}_FY{fy_start_year}-{str(fy_end_year)[-2:]}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Excel: Bank-ready export ──────────────────────────────────────────────────

@router.get("/payroll/export/excel")
async def export_payroll_excel(
    request: Request,
    month_year: str = Query(..., description="Format: YYYY-MM"),
    status: Optional[str] = None,
):
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)

    query: dict = {"month_year": month_year}
    if status:
        query["status"] = status

    records = await db.payroll.find(query, {"_id": 0}).sort("employee_id", 1).to_list(1000)
    if not records:
        raise HTTPException(status_code=404, detail=f"No payroll records for {month_year}.")

    emp_ids = [r["employee_id"] for r in records]
    emps    = await db.employees.find(
        {"employee_id": {"$in": emp_ids}}, {"_id": 0}
    ).to_list(len(emp_ids))
    emp_map = {e["employee_id"]: decrypt_bank_fields(e) for e in emps}

    xlsx_bytes = _generate_excel_export(records, emp_map, month_year)
    filename   = f"Payroll_{month_year}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── PDF: Admin payroll summary export ────────────────────────────────────────

@router.get("/payroll/export/pdf")
async def export_payroll_pdf(
    request: Request,
    month_year: str = Query(..., description="Format: YYYY-MM"),
    status: Optional[str] = None,
):
    await require_roles(UserRole.ADMIN, UserRole.ACCOUNTANT)(request)

    query: dict = {"month_year": month_year}
    if status:
        query["status"] = status

    records = await db.payroll.find(query, {"_id": 0}).sort("employee_id", 1).to_list(1000)
    if not records:
        raise HTTPException(status_code=404, detail=f"No payroll records for {month_year}.")

    emp_ids = [r["employee_id"] for r in records]
    emps    = await db.employees.find(
        {"employee_id": {"$in": emp_ids}}, {"_id": 0}
    ).to_list(len(emp_ids))
    emp_map = {e["employee_id"]: decrypt_bank_fields(e) for e in emps}

    pdf_bytes = _generate_payroll_summary_pdf(records, emp_map, month_year)
    filename  = f"Payroll_Summary_{month_year}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── PDF generation helpers ────────────────────────────────────────────────────

_ORANGE  = colors.HexColor("#E88A1A")
_DARK    = colors.HexColor("#1a1a1a")
_LIGHT   = colors.HexColor("#f5f5f5")
_WHITE   = colors.white
_GREY    = colors.HexColor("#888888")
_GREEN   = colors.HexColor("#2e7d32")


def _base_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("SchoolHeader", parent=styles["Title"],
                              fontSize=16, textColor=_DARK, spaceAfter=2, alignment=TA_CENTER))
    styles.add(ParagraphStyle("SubHeader", parent=styles["Normal"],
                              fontSize=9, textColor=_GREY, alignment=TA_CENTER, spaceAfter=6))
    styles.add(ParagraphStyle("SectionTitle", parent=styles["Normal"],
                              fontSize=10, textColor=_ORANGE, fontName="Helvetica-Bold", spaceAfter=4))
    styles.add(ParagraphStyle("Small", parent=styles["Normal"],
                              fontSize=8, textColor=_GREY))
    styles.add(ParagraphStyle("Bold", parent=styles["Normal"],
                              fontName="Helvetica-Bold", fontSize=10))
    styles.add(ParagraphStyle("Amount", parent=styles["Normal"],
                              fontName="Helvetica-Bold", fontSize=12, textColor=_GREEN, alignment=TA_RIGHT))
    return styles


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _generate_payslip_pdf(record: dict, emp: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=1.5*cm, rightMargin=1.5*cm)
    styles = _base_styles()
    story  = []

    month_name = calendar.month_name[record["month"]]
    emp_name   = f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip()

    # ── Header ──
    story.append(Paragraph(SCHOOL_NAME, styles["SchoolHeader"]))
    story.append(Paragraph(SCHOOL_ADDRESS, styles["SubHeader"]))
    story.append(HRFlowable(width="100%", thickness=2, color=_ORANGE, spaceAfter=8))
    story.append(Paragraph(f"SALARY SLIP — {month_name.upper()} {record['year']}", styles["SectionTitle"]))
    story.append(Spacer(1, 6))

    # ── Employee details ──
    emp_table_data = [
        ["Employee Name", emp_name,         "Employee ID",  emp.get("employee_id", "")],
        ["Designation",   emp.get("designation", "—"),  "Department",   emp.get("department", "—")],
        ["Joining Date",  emp.get("joining_date", "—"), "Status",       record.get("status", "").upper()],
        ["Bank Account",  emp.get("bank_account_number", "—"), "IFSC",  emp.get("bank_ifsc", "—")],
    ]
    emp_table = Table(emp_table_data, colWidths=[3*cm, 6.5*cm, 3*cm, 6*cm])
    emp_table.setStyle(TableStyle([
        ("FONTNAME",    (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 0), (-1, -1), 9),
        ("FONTNAME",    (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME",    (2, 0), (2, -1), "Helvetica-Bold"),
        ("TEXTCOLOR",   (0, 0), (0, -1), _GREY),
        ("TEXTCOLOR",   (2, 0), (2, -1), _GREY),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [_LIGHT, _WHITE]),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("PADDING",     (0, 0), (-1, -1), 5),
    ]))
    story.append(emp_table)
    story.append(Spacer(1, 12))

    # ── Attendance / Day breakdown ──
    story.append(Paragraph("Attendance Summary", styles["SectionTitle"]))
    att_data = [
        ["Total Days in Month", "Working Days", "LWP Days", "Present Days"],
        [str(record["total_days"]), str(record["working_days"]),
         str(record["lwp_days"]), str(record["present_days"])],
    ]
    if record.get("is_mid_month_join"):
        att_data[1].append("(Mid-month join)")
    att_table = Table(att_data, colWidths=[4.5*cm, 4.5*cm, 4.5*cm, 4.5*cm])
    att_table.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), _DARK),
        ("TEXTCOLOR",   (0, 0), (-1, 0), _WHITE),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 9),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("PADDING",     (0, 0), (-1, -1), 6),
    ]))
    story.append(att_table)
    story.append(Spacer(1, 12))

    # ── Earnings / Deductions ──
    story.append(Paragraph("Earnings & Deductions", styles["SectionTitle"]))
    earn_data = [
        ["EARNINGS",            "",         "DEDUCTIONS",           ""],
        ["Monthly Salary",      f"Rs.{record['monthly_salary']:,.2f}",
         "LWP Deduction",       f"Rs.{record['lwp_deduction']:,.2f}"],
        ["Gross Salary",        f"Rs.{record['gross_salary']:,.2f}",
         "Other Deductions",    f"Rs.{record['other_deductions']:,.2f}"],
        ["",                    "",
         "Total Deductions",    f"Rs.{record['total_deductions']:,.2f}"],
    ]
    earn_table = Table(earn_data, colWidths=[4.5*cm, 4.5*cm, 4.5*cm, 4.5*cm])
    earn_table.setStyle(TableStyle([
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 9),
        ("BACKGROUND",  (0, 0), (1, 0), _ORANGE),
        ("BACKGROUND",  (2, 0), (3, 0), _DARK),
        ("TEXTCOLOR",   (0, 0), (-1, 0), _WHITE),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_LIGHT, _WHITE]),
        ("ALIGN",       (1, 0), (1, -1), "RIGHT"),
        ("ALIGN",       (3, 0), (3, -1), "RIGHT"),
        ("PADDING",     (0, 0), (-1, -1), 6),
    ]))
    story.append(earn_table)
    story.append(Spacer(1, 14))

    # ── Net Salary ──
    net_data = [["NET SALARY PAYABLE", f"Rs.{record['net_salary']:,.2f}"]]
    net_table = Table(net_data, colWidths=[13*cm, 5.5*cm])
    net_table.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, -1), _GREEN),
        ("TEXTCOLOR",   (0, 0), (-1, -1), _WHITE),
        ("FONTNAME",    (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 13),
        ("ALIGN",       (1, 0), (1, 0), "RIGHT"),
        ("PADDING",     (0, 0), (-1, -1), 10),
    ]))
    story.append(net_table)
    story.append(Spacer(1, 16))

    if record.get("deduction_remarks"):
        story.append(Paragraph(f"Remarks: {record['deduction_remarks']}", styles["Small"]))
    story.append(Spacer(1, 10))
    story.append(Paragraph("This is a computer-generated payslip and does not require a signature.",
                            styles["Small"]))

    doc.build(story)
    return buf.getvalue()


def _generate_yearly_statement_pdf(records: list, emp: dict, year: int) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=1.5*cm, rightMargin=1.5*cm)
    styles = _base_styles()
    story  = []

    emp_name = f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip()

    story.append(Paragraph(SCHOOL_NAME, styles["SchoolHeader"]))
    story.append(Paragraph(SCHOOL_ADDRESS, styles["SubHeader"]))
    story.append(HRFlowable(width="100%", thickness=2, color=_ORANGE, spaceAfter=8))
    story.append(Paragraph(f"ANNUAL SALARY STATEMENT — {year}", styles["SectionTitle"]))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"Employee: {emp_name}  |  ID: {emp.get('employee_id', '')}  |  Designation: {emp.get('designation', '')}",
                            styles["SubHeader"]))
    story.append(Spacer(1, 10))

    headers = ["Month", "Gross Salary", "LWP Days", "LWP Deduct.",
               "Other Deduct.", "Total Deduct.", "Net Salary", "Status"]
    rows = [headers]
    total_gross = total_net = total_lwp_ded = total_other_ded = 0.0

    for r in records:
        rows.append([
            calendar.month_abbr[r["month"]],
            f"Rs.{r['gross_salary']:,.2f}",
            str(r["lwp_days"]),
            f"Rs.{r['lwp_deduction']:,.2f}",
            f"Rs.{r['other_deductions']:,.2f}",
            f"Rs.{r['total_deductions']:,.2f}",
            f"Rs.{r['net_salary']:,.2f}",
            r["status"].upper(),
        ])
        total_gross    += r["gross_salary"]
        total_net      += r["net_salary"]
        total_lwp_ded  += r["lwp_deduction"]
        total_other_ded += r["other_deductions"]

    rows.append([
        "TOTAL",
        f"Rs.{total_gross:,.2f}", "", f"Rs.{total_lwp_ded:,.2f}",
        f"Rs.{total_other_ded:,.2f}", f"Rs.{total_lwp_ded+total_other_ded:,.2f}",
        f"Rs.{total_net:,.2f}", "",
    ])

    col_widths = [2*cm, 3*cm, 2*cm, 2.8*cm, 2.8*cm, 2.8*cm, 3*cm, 2*cm]
    tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), _DARK),
        ("TEXTCOLOR",     (0, 0), (-1, 0), _WHITE),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 8),
        ("ALIGN",         (1, 0), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [_WHITE, _LIGHT]),
        ("BACKGROUND",    (0, -1), (-1, -1), _ORANGE),
        ("TEXTCOLOR",     (0, -1), (-1, -1), _WHITE),
        ("FONTNAME",      (0, -1), (-1, -1), "Helvetica-Bold"),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("PADDING",       (0, 0), (-1, -1), 5),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 16))
    story.append(Paragraph("This is a computer-generated statement.", styles["Small"]))

    doc.build(story)
    return buf.getvalue()


def _generate_form16_pdf(records: list, emp: dict, fy_start: int, fy_end: int) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=1.5*cm, rightMargin=1.5*cm)
    styles = _base_styles()
    story  = []

    emp_name = f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip()

    story.append(Paragraph(SCHOOL_NAME, styles["SchoolHeader"]))
    story.append(Paragraph(SCHOOL_ADDRESS, styles["SubHeader"]))
    story.append(HRFlowable(width="100%", thickness=2, color=_ORANGE, spaceAfter=8))
    story.append(Paragraph(
        f"FORM 16 — CERTIFICATE OF SALARY (FY {fy_start}-{str(fy_end)[-2:]})",
        styles["SectionTitle"]
    ))
    story.append(Spacer(1, 6))

    total_gross = sum(r.get("gross_salary", 0) for r in records)
    total_deductions = sum(r.get("total_deductions", 0) for r in records)
    total_net = sum(r.get("net_salary", 0) for r in records)

    emp_info = [
        ["Name of Employee",        emp_name],
        ["Employee ID",             emp.get("employee_id", "")],
        ["Designation",             emp.get("designation", "—")],
        ["Department",              emp.get("department", "—")],
        ["Financial Year",          f"April {fy_start} to March {fy_end}"],
    ]
    info_tbl = Table(emp_info, colWidths=[6*cm, 12.5*cm])
    info_tbl.setStyle(TableStyle([
        ("FONTNAME",    (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 9),
        ("TEXTCOLOR",   (0, 0), (0, -1), _GREY),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [_LIGHT, _WHITE]),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("PADDING",     (0, 0), (-1, -1), 5),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 12))

    story.append(Paragraph("Part A — Salary Details", styles["SectionTitle"]))
    salary_data = [
        ["Description",                        "Amount (Rs.)"],
        ["Gross Salary Paid",                  f"Rs.{total_gross:,.2f}"],
        ["Total Deductions (LWP + Other)",     f"Rs.{total_deductions:,.2f}"],
        ["Net Salary Paid",                    f"Rs.{total_net:,.2f}"],
        ["Tax Deducted at Source (TDS)",       "Rs.0.00  (as applicable)"],
    ]
    sal_tbl = Table(salary_data, colWidths=[13*cm, 5.5*cm])
    sal_tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), _DARK),
        ("TEXTCOLOR",   (0, 0), (-1, 0), _WHITE),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_WHITE, _LIGHT]),
        ("ALIGN",       (1, 0), (1, -1), "RIGHT"),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("PADDING",     (0, 0), (-1, -1), 6),
    ]))
    story.append(sal_tbl)
    story.append(Spacer(1, 12))

    story.append(Paragraph("Part B — Month-wise Breakdown", styles["SectionTitle"]))
    month_headers = ["Month", "Gross", "Deductions", "Net Salary", "Status"]
    month_rows    = [month_headers]
    for r in records:
        month_rows.append([
            f"{calendar.month_abbr[r['month']]} {r['year']}",
            f"Rs.{r['gross_salary']:,.2f}",
            f"Rs.{r['total_deductions']:,.2f}",
            f"Rs.{r['net_salary']:,.2f}",
            r["status"].upper(),
        ])
    month_tbl = Table(month_rows, colWidths=[4*cm, 4*cm, 4*cm, 4*cm, 2.5*cm], repeatRows=1)
    month_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), _ORANGE),
        ("TEXTCOLOR",     (0, 0), (-1, 0), _WHITE),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 8),
        ("ALIGN",         (1, 0), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_WHITE, _LIGHT]),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("PADDING",       (0, 0), (-1, -1), 5),
    ]))
    story.append(month_tbl)
    story.append(Spacer(1, 20))
    story.append(Paragraph(
        "Certified that the above information is correct to the best of our knowledge.",
        styles["Normal"]
    ))
    story.append(Spacer(1, 30))
    story.append(Paragraph("Authorised Signatory", styles["Bold"]))
    story.append(Paragraph(SCHOOL_NAME, styles["Small"]))
    story.append(Spacer(1, 10))
    story.append(Paragraph("Note: This is a computer-generated Form 16 for reference. "
                            "For TDS certificate purposes, consult your CA.", styles["Small"]))

    doc.build(story)
    return buf.getvalue()


def _generate_payroll_summary_pdf(records: list, emp_map: dict, month_year: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=1*cm, rightMargin=1*cm)
    styles = _base_styles()
    story  = []

    story.append(Paragraph(SCHOOL_NAME, styles["SchoolHeader"]))
    story.append(Paragraph(SCHOOL_ADDRESS, styles["SubHeader"]))
    story.append(HRFlowable(width="100%", thickness=2, color=_ORANGE, spaceAfter=8))
    story.append(Paragraph(f"PAYROLL SUMMARY — {month_year}", styles["SectionTitle"]))
    story.append(Spacer(1, 6))

    headers = ["#", "Name", "Designation", "Gross (Rs.)", "Deduct. (Rs.)", "Net (Rs.)", "Status"]
    rows = [headers]
    total_gross = total_net = total_ded = 0.0

    for i, r in enumerate(records, 1):
        emp = emp_map.get(r["employee_id"], {})
        name = f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip()
        rows.append([
            str(i), name, emp.get("designation", "—"),
            f"{r['gross_salary']:,.2f}",
            f"{r['total_deductions']:,.2f}",
            f"{r['net_salary']:,.2f}",
            r["status"].upper(),
        ])
        total_gross += r["gross_salary"]
        total_ded   += r["total_deductions"]
        total_net   += r["net_salary"]

    rows.append(["", "TOTAL", "", f"{total_gross:,.2f}", f"{total_ded:,.2f}", f"{total_net:,.2f}", ""])

    col_widths = [0.8*cm, 5*cm, 3.5*cm, 3*cm, 3*cm, 3*cm, 2.2*cm]
    tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), _DARK),
        ("TEXTCOLOR",     (0, 0), (-1, 0), _WHITE),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 8),
        ("ALIGN",         (3, 0), (5, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [_WHITE, _LIGHT]),
        ("BACKGROUND",    (0, -1), (-1, -1), _ORANGE),
        ("TEXTCOLOR",     (0, -1), (-1, -1), _WHITE),
        ("FONTNAME",      (0, -1), (-1, -1), "Helvetica-Bold"),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("PADDING",       (0, 0), (-1, -1), 5),
    ]))
    story.append(tbl)

    doc.build(story)
    return buf.getvalue()


# ── Excel generation helper ───────────────────────────────────────────────────

def _generate_excel_export(records: list, emp_map: dict, month_year: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Payroll {month_year}"

    orange_fill = PatternFill("solid", fgColor="E88A1A")
    dark_fill   = PatternFill("solid", fgColor="1A1A1A")
    light_fill  = PatternFill("solid", fgColor="F5F5F5")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    bold_font   = Font(bold=True, size=10)
    normal_font = Font(size=9)
    center      = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right",  vertical="center")

    def thin_border_cell(cell):
        s = Side(style="thin", color="CCCCCC")
        cell.border = Border(left=s, right=s, top=s, bottom=s)

    # ── Title row ──
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value     = f"{SCHOOL_NAME} — Payroll Register {month_year}"
    title_cell.font      = Font(bold=True, size=13, color="1A1A1A")
    title_cell.alignment = Alignment(horizontal="center")
    title_cell.fill      = PatternFill("solid", fgColor="FFF3E0")
    ws.row_dimensions[1].height = 28

    ws.append([])   # blank row

    # ── Header row ──
    headers = [
        "#", "Employee ID", "Name", "Designation", "Department",
        "Phone", "Address",
        "Bank Account", "IFSC", "Bank Name",
        "Gross Salary (Rs.)", "Deductions (Rs.)", "Net Salary (Rs.)", "Status"
    ]
    ws.append(headers)
    header_row = ws.max_row
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font      = header_font
        cell.fill      = dark_fill
        cell.alignment = center
        thin_border_cell(cell)
    ws.row_dimensions[header_row].height = 20

    # ── Data rows ──
    total_gross = total_ded = total_net = 0.0
    for i, r in enumerate(records, 1):
        emp  = emp_map.get(r["employee_id"], {})
        name = f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip()
        row  = [
            i,
            r["employee_id"],
            name,
            emp.get("designation", ""),
            emp.get("department", ""),
            emp.get("phone", ""),
            emp.get("address", ""),
            emp.get("bank_account_number", r.get("bank_account_number", "")),
            emp.get("bank_ifsc",           r.get("bank_ifsc", "")),
            emp.get("bank_name",           r.get("bank_name", "")),
            round(r["gross_salary"], 2),
            round(r["total_deductions"], 2),
            round(r["net_salary"], 2),
            r["status"].upper(),
        ]
        ws.append(row)
        data_row = ws.max_row
        fill = light_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=data_row, column=col_num)
            cell.font  = normal_font
            cell.fill  = fill
            thin_border_cell(cell)
            if col_num in (11, 12, 13):
                cell.alignment  = right_align
                cell.number_format = '#,##0.00'
            else:
                cell.alignment  = Alignment(vertical="center")

        total_gross += r["gross_salary"]
        total_ded   += r["total_deductions"]
        total_net   += r["net_salary"]

    # ── Totals row ──
    ws.append([])
    totals_row_num = ws.max_row + 1
    totals = ["", "", "TOTAL", "", "", "", "", "", "", "",
              round(total_gross, 2), round(total_ded, 2), round(total_net, 2), ""]
    ws.append(totals)
    for col_num, val in enumerate(totals, 1):
        cell = ws.cell(row=totals_row_num, column=col_num)
        cell.font  = bold_font
        cell.fill  = orange_fill
        thin_border_cell(cell)
        if col_num in (11, 12, 13):
            cell.alignment     = right_align
            cell.number_format = '#,##0.00'
        if col_num == 3:
            cell.font = Font(bold=True, size=10, color="FFFFFF")

    # ── Column widths ──
    col_widths = [4, 14, 22, 18, 14, 14, 28, 20, 14, 16, 16, 16, 16, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Bank transfer sheet (clean, minimal) ──
    ws2 = wb.create_sheet("Bank Transfer")
    bt_headers = ["#", "Name", "Account Number", "IFSC", "Bank Name", "Amount (Rs.)", "Remarks"]
    ws2.append(bt_headers)
    for col_num, h in enumerate(bt_headers, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = orange_fill
        cell.alignment = center
        thin_border_cell(cell)

    for i, r in enumerate(records, 1):
        emp  = emp_map.get(r["employee_id"], {})
        name = f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip()
        row  = [
            i, name,
            emp.get("bank_account_number", r.get("bank_account_number", "")),
            emp.get("bank_ifsc",           r.get("bank_ifsc", "")),
            emp.get("bank_name",           r.get("bank_name", "")),
            round(r["net_salary"], 2),
            f"Salary {month_year}",
        ]
        ws2.append(row)
        data_row = ws2.max_row
        for col_num in range(1, 8):
            cell = ws2.cell(row=data_row, column=col_num)
            cell.font = normal_font
            thin_border_cell(cell)
            if col_num == 6:
                cell.alignment     = right_align
                cell.number_format = '#,##0.00'

    for i, w in enumerate([4, 22, 20, 14, 16, 16, 18], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
